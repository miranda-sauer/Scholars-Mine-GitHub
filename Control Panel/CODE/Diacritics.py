# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           IMPORT                                                                  #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import time
import os
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook
import datetime



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           DISPLAY WINDOW                                                          #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def open_diacritics():
    # Create window
    diacritics = Toplevel()
    diacritics.title("Diacritics Program")
    diacritics.configure(bg = '#003B49')
    diacritics.iconbitmap('R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Control Panel/CODE/S&T_Logo.ico')

    # Center the window on the screen
    window_w = 400 # window width
    window_h = 290 # window height

    screen_w = diacritics.winfo_screenwidth() # screen width
    screen_h = diacritics.winfo_screenheight() # screen height

    x_coor = (screen_w/2) - (window_w/2) #middle of screen x coordinate
    y_coor = (screen_h/2) #lower half of screen y coordinate

    diacritics.geometry("%dx%d+%d+%d" % (window_w, window_h, x_coor, y_coor)) # place in middle

    #Create Progress Bar
    progress = ttk.Progressbar(diacritics, orient = HORIZONTAL, length = 370, mode = 'determinate')

    # Create a frame
    frame = LabelFrame(diacritics, text = "Select an Input File", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')

    # Create a label
    task = Label(diacritics, text = "Waiting for a file", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')
    file_label = Label(diacritics, text = "Folder: N/A", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')


    # Open Help Function
    def open_help():
        # Open word document
        os.startfile("R:/storage/libarchive/b/1. Processing/8. Other Projects/Control Panel/Documentation/diacritics Help.docx")

    # Create a button
    help_button = Button(diacritics, text = "Help", command = lambda : open_help(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    exit_button = Button(diacritics, text = "Exit Diacritics", command = diacritics.destroy, bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")

    # Place in window
    frame.pack(padx = 20, pady = 10)
    file_label.pack(padx = 20, pady = (0, 20))
    progress.pack(padx = 5, pady = 5)
    task.pack(padx = 20)
    help_button.pack(padx = (331, 10), pady = 0)
    exit_button.pack(padx = (262, 10), pady = (5, 0))

    # Error Message Popup
    def error_popup(error_message):
        messagebox.showerror("Error", error_message)

    # Warning Message Popup
    def warning_popup(warning_message):
        messagebox.showwarning("Warning", warning_message)

    # Update Progress Bar
    def update_progress(p, t):  
        # Update bar value
        progress['value'] = (p/3)*100

        # Update bar label
        task.config(text = t)
        task.pack() 

        # Refresh window (very important line!)
        diacritics.update_idletasks()
        time.sleep(0.2)

    # Function Click
    def browse():
        # Reset progress bar
        update_progress(0, "Waiting for a file")

        # Open file
        diacritics.filename = filedialog.askopenfilename(initialdir = "R:/storage/libarchive", title = "Select Input", filetypes = (("Excel Workbook", "*.xlsx"),))
        if diacritics.filename == "":
            warning_popup("No folder selected.")
            file_label.config(text = "Folder: N/A")
            del diacritics.filename
        else:
            #Get file name
            name = diacritics.filename
            for i in range(len(name)):
                if "/" in name[-(i)]:
                    name = "Folder: " + name[-(i-1):]
                    break

            #Update Folder Name Label
            file_label.config(text = name)

    # Main Function
    def main(file_name):
        update_progress(1, "Fixing diacritics...")
        
        # Load Excel File
        try:
            path = str(file_name)
            book = load_workbook(path)  #read file
            sheet = book.active         #read sheet
        except:
            error_popup("Couldn't Load Excel")

        # Determine max row and column
        try:
            max_col = len([c for c in sheet.iter_cols(min_row=1, max_row=1, values_only=True) if c[0] is not None])
            max_row = len([c for c in sheet.iter_rows(min_col=1, max_col=1, values_only=True) if c[0] is not None])
            print("Rows: " + str(max_row))
            print("Columns: " + str(max_col))
        except:
            error_popup("Couldn't Determine Max Row/Col")

        # Search Excel
        try:
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    if sheet.cell(row, col).value == "kb":
                        sheet.cell(row, col).value = "fixed"
        except:
            error_popup("Couldn't Search File")
        
        update_progress(2, "Saving file...") # Saving

        # Save excel
        try:
            path = str(file_name)[:len(file_name) - 5] + '_Complete.xlsx'
            book.save(path)
        except:
            error_popup("Couldn't Save File")

        update_progress(3, "Excel Created")
    
    # Start Button Function
    def start():
        # Run program and update progress bar
        try:
            main(diacritics.filename)
        except AttributeError:
            error_popup("No file selected. Browse to select a file.")
        except:
            error_popup("There was an unknown error, the file could not be processed.")
    
    # Create a button
    browse_button = Button(frame, text = "Browse", command = lambda : browse(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    start_button = Button(frame, text = "Start", command = lambda : start(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")

    # Place in frame
    browse_button.grid(row = 0, column = 0, padx = (15, 0), pady = 15)
    start_button.grid(row = 0, column = 1, padx = 15, pady = 15)