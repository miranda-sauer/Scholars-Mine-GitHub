# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           IMPORT                                                                  #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
import time
import os
import shutil
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           DISPLAY WINDOW                                                          #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def open_reformat_harvesting():
    # Create window
    reformat_harvesting = Toplevel()
    reformat_harvesting.title("Reformat Harvesting Program")
    reformat_harvesting.configure(bg = '#003B49')
    reformat_harvesting.iconbitmap('R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Control Panel/CODE/S&T_Logo.ico')

    # Center the window on the screen
    window_w = 400 # window width
    window_h = 290 # window height

    screen_w = reformat_harvesting.winfo_screenwidth() # screen width
    screen_h = reformat_harvesting.winfo_screenheight() # screen height

    x_coor = (screen_w/2) - (window_w/2) #middle of screen x coordinate
    y_coor = (screen_h/2) #lower half of screen y coordinate

    reformat_harvesting.geometry("%dx%d+%d+%d" % (window_w, window_h, x_coor, y_coor)) # place in middle

    #Create Progress Bar
    progress = ttk.Progressbar(reformat_harvesting, orient = HORIZONTAL, length = 370, mode = 'determinate')

    # Create a frame
    frame = LabelFrame(reformat_harvesting, text = "Select an Input File", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')

    # Create a label
    task = Label(reformat_harvesting, text = "Waiting for a file", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')
    file_label = Label(reformat_harvesting, text = "Folder: N/A", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')


    # Open Help Function
    def open_help():
        # Open word document
        os.startfile("R:/storage/libarchive/b/1. Processing/8. Other Projects/Control Panel/Documentation/reformat_harvesting Help.docx")

    # Create a button
    help_button = Button(reformat_harvesting, text = "Help", command = lambda : open_help(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    exit_button = Button(reformat_harvesting, text = "Exit Reformat Harvesting", command = reformat_harvesting.destroy, bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")

    # Place in window
    frame.pack(padx = 20, pady = 10)
    file_label.pack(padx = 20, pady = (0, 20))
    progress.pack(padx = 5, pady = 5)
    task.pack(padx = 20)
    help_button.pack(padx = (331, 10), pady = 0)
    exit_button.pack(padx = (175, 10), pady = (5, 0))

    # Update Progress Bar
    def update_progress(p, t):  
        # Update bar value
        progress['value'] = (p/3)*100

        # Update bar label
        task.config(text = t)
        task.pack() 

        # Refresh window (very important line!)
        reformat_harvesting.update_idletasks()
        time.sleep(0.2)

    # Function Click
    def browse():
        # Reset progress bar
        update_progress(0, "Waiting for a file")

        # Open file
        reformat_harvesting.filename = filedialog.askopenfilename(initialdir = "R:/storage/libarchive/b/1. Processing/8. Other Projects/Harvesting/Excel Files", title = "Select Input", filetypes = (("Excel Workbook", "*.xlsx"),))

        #Get file name
        name = reformat_harvesting.filename
        for i in range(len(name)):
            if "/" in name[-(i)]:
                name = "Folder: " + name[-(i-1):]
                break

        #Update Folder Name Label
        file_label.config(text = name)

    # Main Function
    def main(file_name):
        update_progress(1, "Reformating Harvesting Excel File...")

        # Load Excel File
        path = str(file_name)
        book = load_workbook(path)  #read file
        sheet = book.active         #read sheet

        # Determine how many columns are in the sheet that is being read from
        max_col = len([c for c in sheet.iter_cols(min_row=1, max_row=1, values_only=True) if c[0] is not None])

        # Starting author count
        last_cell = str(sheet.cell(1, max_col).value)[6:]

        index = 0
        for letter in last_cell:
            if letter == '_':
                author_count = int(last_cell[:index])
                break
            index += 1

        def index(header):
            for col in range(1, max_col + 1):
                col_head = sheet.cell(1, col).value
                if header in col_head:
                    return col

        def specific_index(header, do_not_include):
            for col in range(1, max_col + 1):
                col_head = sheet.cell(1, col).value
                if header in col_head and do_not_include not in col_head:
                    return col

        # Copy Information
        # new_col is the column index where info is being copied to
        # old_col is the column index where info is being copied from
        def copy(new_col, old_col):
            sheet.cell(row, new_col).value = sheet.cell(row, old_col).value

        # Fill Information
        # col is the column index where info is being filled (new excel)
        # fill_text is the text that is filled into the new excel
        def fill(col, fill_text):
            sheet.cell(row, col).value = fill_text

        # Fill author_split column
        def author_split_information(row):

            for num in range(1, int(sheet.cell(row, index("total_author_count")).value) + 1):
                # name information
                last_name = sheet.cell(row, index("author" + str(num) + "_lname")).value
                first_name = sheet.cell(row, index("author" + str(num) + "_fname")).value
                
                if sheet.cell(row, index("author" + str(num) + "_mname")).value:
                    middle_name = sheet.cell(row, index("author" + str(num) + "_mname")).value
                    middle_initial = middle_name[0:1]

                    if num > 1:
                        author_split = str(sheet.cell(row, specific_index("author", "_")).value) + " and " + str(last_name) + ", " + str(first_name) + " " + str(middle_initial) + "."
                    else:
                        author_split = str(last_name) + ", " + str(first_name) + " " + str(middle_initial) + "."
                else:
                    if num > 1:
                        author_split = str(sheet.cell(row, specific_index("author", "_")).value) + " and " + str(last_name) + ", " + str(first_name)
                    else:
                        author_split = str(last_name) + ", " + str(first_name)

                try:
                    # author_split
                    fill(specific_index("author", "_"), str(author_split))
                except TypeError:
                    if row == 2:
                        print("Couldn't find 'author' column.")


        # Determine how many rows are in the sheet that is being read from
        max_row = len([c for c in sheet.iter_rows(min_col=1, max_col=1, values_only=True) if c[0] is not None])
        #max_row -= 1 # Because of the "in Scholars' Mine" at the bottom

        # Transfer row information for the whole sheet
        update_progress(2, "Reformating Harvesting...") # Filling in " + str(max_row) + " rows

        for row in range(2, max_row + 1):
            author_split_information(row)

        # Save excel
        update_progress(3, "Saving file...") # Saving

        path = str(file_name)[:len(file_name) - 5] + '_Reformatted.xlsx'
        book.save(path)

        update_progress(3, "Excel Created")
    
    # Start Button Function
    def start():
        # Run program and update progress bar
        main(reformat_harvesting.filename)
    
    # Create a button
    browse_button = Button(frame, text = "Browse", command = lambda : browse(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    start_button = Button(frame, text = "Start", command = lambda : start(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")

    # Place in frame
    browse_button.grid(row = 0, column = 0, padx = (15, 0), pady = 15)
    start_button.grid(row = 0, column = 1, padx = 15, pady = 15)