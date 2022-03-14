# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#                                       Import                                      #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.colors import Color
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime
import openpyxl as xl
import time
import os
import shutil
import sqlite3
import pickle
import sys
import pandas as pd
import numpy as np
from __init__ import fix_text
_global = sys.modules[__name__] # Allows access to 'global' variables defined below
from author_diacritics import ensure_encryption
from dateutil.relativedelta import relativedelta
import re
from author_diacritics import ensure_encryption


def open_both():
    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
    #                      Variable Declaration and Initialization                      #
    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

    Both = Toplevel() # Create window
    top_frame = Frame(Both, bg = '#003B49') # Create a frame
    help_button = Button(top_frame, text = "Help", command = lambda : open_help_documentation(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge") # Create a button
    select_button = Button(top_frame, text = "Select File(s)", command = lambda : browse(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge") # Create a button

    middle_frame = Frame(Both, bg = '#003B49') # Create a frame
    label_canvas = Canvas(middle_frame, bg = '#003B49') # Create a canvas
    canvas_frame = Frame(label_canvas, bg = '#003B49') # Create a frame
    file_label = Label(canvas_frame, text = "n/a", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold', anchor = 'w', justify = 'left') # Create a label
    scroll_bar = ttk.Scrollbar(middle_frame, orient = VERTICAL, command = label_canvas.yview)

    bottom_frame = Frame(Both, bg = '#003B49') # Create a frame
    progress = ttk.Progressbar(bottom_frame, orient = HORIZONTAL, length = 400, mode = 'determinate') # Create Progress Bar
    start_button = Button(bottom_frame, text = "Start", command = lambda : start(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge") # Create a button

    filenames = tuple()



    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
    #                      Function Declaration and Initialization                      #
    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
    # Open Help Documentation
    def open_help_documentation():
        os.startfile('R:/storage/libarchive/a/zzz_Programs/Control Panel/Documentation/Excel Processing Help.pdf')


    # Error Message Popup
    def error_popup(error_message):
        messagebox.showerror("Error", error_message)


    # Warning Message Popup
    def warning_popup(warning_message):
        messagebox.showwarning("Warning", warning_message)


    # Information Message Popup
    def info_popup(info_message):
        messagebox.showinfo('Information', info_message)


    # Update Progress Bar
    def update_progress(p):  
        # Update bar value
        progress['value'] = (p/10)*100

        # Refresh window (very important line!)
        Both.update_idletasks()
        time.sleep(0.2)


    # Select File(s) Button Function
    def browse():
        # Reset progress bar
        update_progress(0)

        # Open file
        global filenames
        filenames = filedialog.askopenfilename(title = "Select Input", initialdir = "R:/storage/libarchive/b/1. Processing/8. Other Projects/Excel Files", filetypes = (("Excel Workbook", "*.xlsx"),), multiple = True)
        if len(filenames) == 0:
            warning_popup("No file(s) selected.")
            file_label.config(text = "n/a")
        else:
            #Update File Name Label
            name_list = ''
            for index in range(len(filenames)):
                last_index = filenames[index].rfind('/')
                name_list = name_list + str(filenames[index])[last_index + 1:]
                if index != (len(filenames) - 1):
                    name_list = name_list + '\n'

            file_label.config(text = name_list)


    # Start Button Function
    def start():
        global filenames
        
        for index in range(len(filenames)):
            print('Processing: ' + str(filenames[index]))
            author_split_main(filenames[index])
            harvesting_main(filenames[index])
        

    # Author Split Main Function
    def author_split_main(file_name):
        update_progress(1)
                
        rdsheet = None
        author_column = ''
        excelName = ''

        authority_database = sqlite3.connect('R:/storage/libarchive/a/zzz_Programs/faculty.db')
        authority_cursor = authority_database.cursor()
        def regexp(expr, item):
            reg = re.compile(expr)
            return reg.search(item) is not None
        authority_database.create_function("REGEXP", 2, regexp)

        sqlite3.enable_callback_tracebacks(True)   # <-- !\

        authorDict = {}
        rb = None

        # Read in Excel file
        with pd.ExcelFile(file_name) as excel_in:
            df = pd.read_excel( excel_in, excel_in.sheet_names[0], header=0, index_col=False )

        for i, c in enumerate( df.columns ):
            if "_fname" in c:
                df.iloc[ :, i ] = df.iloc[ :, i ].fillna('')
                df.iloc[ :, i + 1 ] = df.iloc[ :, i + 1 ].fillna('')
                for j in range( df.shape[0] ):
                    split_name = df.iloc[ j, i ].split(" ")
                    df.iloc[ j, i ] = split_name[0]
                    split_name = split_name[1:]
                    df.iloc[ j, i + 1 ] = " ".join( split_name )

        # Find columns corresponding to S&T affiliates, that is, columns whose
        ## "autho<i>_institution" is "Missouri University of Science and Technology"
        author_columns = []
        for r in range( df.shape[0] ):
            for i, c in enumerate(df.columns):
                if "_institution" in c and df.loc[r, c] == "Missouri University of Science and Technology":
                    author_columns.append( (r, i-5) )

        for c in df.columns:
            if "_institution" in c:
                # print(c, i+1)
                t_df = df.loc[:, c] = ""

        # $author_columns contains S&T graduate students, to filter them out, we will
        ## filter them out by looking them up in the faculty.db
        df["faculty_author_count"] = np.zeros( df.shape[0] )
        dct_df = {}
        for index in author_columns:
            t_df = df.iloc[ index[0], index[1]:index[1]+7 ]

            t_name_dct = { "first": t_df.values[0].split(" ")[0]
                            , "last":t_df.values[2].split(" ")[0]
                        }

            t_query = authority_cursor.execute('SELECT authority_name,first_name, middle_name, last_name, email, department FROM faculty WHERE (last_name COLLATE NOCASE) LIKE :last AND (first_name COLLATE NOCASE) LIKE :first', t_name_dct).fetchall()
            if len(t_query) == 1:
                dct_df[ index ] = list(t_query[0])
                df.loc[ index[0], "faculty_author_count" ] += 1
            elif len(t_query) > 1:
                t_query = list(t_query[0])
                for i in [0,2,4,5]:
                    t_query[i] = ""
                t_query[-1] = "MANUAL-CHECK"
                dct_df[ index ] = t_query
                df.loc[ index[0], "faculty_author_count" ] = np.nan
            else:
                for key in t_name_dct:
                    t_name_dct[key] = t_name_dct[key].replace(".","").replace(",","")
                t_query = authority_cursor.execute('SELECT authority_name, first_name, middle_name, last_name, email, department FROM faculty WHERE (last_name COLLATE NOCASE) LIKE :last AND first_name REGEXP :first', t_name_dct).fetchall()
                if len(t_query) == 1:
                    dct_df[ index ] = list(t_query[0])
                    df.loc[ index[0], "faculty_author_count" ] += 1
                elif len(t_query) > 1:
                    t_query = list(t_query[0])
                    for i in [0,2,4,5]:
                        t_query[i] = ""
                    t_query[-1] = "MANUAL-CHECK"
                    dct_df[ index ] = t_query
                    df.loc[ index[0], "faculty_author_count" ] = np.nan

        '''
            For each S&T author, fill in
                author{i}_fname	author{i}_mname	author{i}_lname	author{i}_suffix	author{i}_email	author{i}_institution	author{i}_is_corporate
            with data taken from the faculty database and setting the institution to "Missouri University of Science and Technology"
        '''
        for index in dct_df:
            try:
                df.iloc[ index[0], index[1]:index[1]+7 ] = dct_df[index][1:4] + [""] + dct_df[index][4:5] + ["Missouri University of Science and Technology",""]
            except:
                print(dct_df[index], end="\n\n\n")

        '''
            Get total author count by counting non-empty last name columns
        '''
        df = df.fillna('')
        df["total_author_count"] = np.zeros(df.shape[0])
        for c in df.columns:
            for r in range( df.shape[0] ):
                if "_lname" == c[-6:] and df.loc[ r, c ]:
                    # print(str(df.loc[ r, c ]) )
                    df.loc[r, "total_author_count"] += 1

        '''
            (1) Get authorized names separated by '<br>'
            (2) Get departments 1-4, non-repeating, in order of authors
        '''
        df["authorized_name"] = ""
        for i in range(4):
            df[ f"department{i+1}" ] = ""
        for r in range( df.shape[0] ):
            department_num = 1
            authorized_name_list = ""
            for index in dct_df:
                if index[0] == r:
                    authorized_name_list += dct_df[index][0] + "<br>"
                    if department_num <= 4:
                        department = dct_df[index][-1]
                        department_runner = np.max( [1, department_num-1] )
                        while department_runner >= 1:
                            if department == df.loc[r, f"department{department_runner}"]:
                                break
                            department_runner -= 1
                        # print( department_runner )
                        if department_runner == 0:
                            # print(department)
                            df.loc[r, f"department{department_num}"] = department
                            department_num += 1
            authorized_name_list = authorized_name_list[:-4]
            df.loc[ r, "authorized_name" ] = authorized_name_list

        '''
            Replace diacritics with appropriate token in columns: Abstract, Keywords, Funding Sponsor, Publisher, First name, Middle name, Last names, Source Publications
        '''
        df = ensure_encryption(df)

        # Write in Excel file
        with pd.ExcelWriter(f"{file_name[:-5]}_Author_Split.xlsx") as excel_out:
            df.to_excel(excel_out, sheet_name="Sheet1", header=True, index=False)
            #excel_out.close()
            del df

        update_progress(2)


    # Character to subscript
    def char_to_sub(argument):
        output = ''
        switcher = {
            '0': '\u2080',
            '1': '\u2081',
            '2': '\u2082',
            '3': '\u2083',
            '4': '\u2084',
            '5': '\u2085',
            '6': '\u2086',
            '7': '\u2087',
            '8': '\u2088',
            '9': '\u2089',
            '+': '\u208A',
            '-': '\u208B',
            '=': '\u208C',
            '(': '\u208D',
            ')': '\u208E',
            'a': '\u2090',
            'e': '\u2091',
            'o': '\u2092',
            'x': '\u2093',
            'h': '\u2095',
            'k': '\u2096',
            'l': '\u2097',
            'm': '\u2098',
            'n': '\u2099',
            'p': '\u209A',
            's': '\u209B',
            't': '\u209C',
        }
        output = switcher.get(argument, '?')
        if '?' in output:
            SUB = str.maketrans('aehijklmnoprstuvxy', 'ₐₑₕᵢⱼₖₗₘₙₒₚᵣₛₜᵤᵥₓᵧ')
            output = argument.translate(SUB)
        return output


    # Character to superscript
    def char_to_sup(argument):
        output = ''
        switcher = {
            '0': '\u2070',
            '1': '\u00B9',
            '2': '\u00B2',
            '3': '\u00B3',
            '4': '\u2074',
            '5': '\u2075',
            '6': '\u2076',
            '7': '\u2077',
            '8': '\u2078',
            '9': '\u2079',
            'i': '\u2071',
            '+': '\u207A',
            '-': '\u207B',
            '=': '\u207C',
            '(': '\u207D',
            ')': '\u207E',
            'n': '\u207F',
        }
        output = switcher.get(argument, '?')
        if '?' in output:
            SUP = str.maketrans('ABDEGHIJKLMNOPRTUVWabcdefghijklmnoprstuvwxyz', 'ᴬᴮᴰᴱᴳᴴᴵᴶᴷᴸᴹᴺᴼᴾᴿᵀᵁⱽᵂᵃᵇᶜᵈᵉᶠᵍʰⁱʲᵏˡᵐⁿᵒᵖʳˢᵗᵘᵛʷˣʸᶻ')
            output = argument.translate(SUP)
        return output


    # Fix subscript and superscript
    def replace_sub_sup(title):
        while('</sub>' in title):
            start = title.find('<sub>')
            end = title.find('</sub>')
            title = title.replace('<sub>', '')
            while start != end:
                title = str(title[0:start]) + str(char_to_sub(title[start])) + str(title[start + 1:])
                start+=1
                end = title.find('</sub>')
            title = title.replace('</sub>', '')

        while('</sup>' in title):
            start = title.find('<sup>')
            end = title.find('</sup>')
            title = title.replace('<sup>', '')
            while start != end:
                title = str(title[0:start]) + str(char_to_sup(title[start])) + str(title[start + 1:])
                start+=1
                end = title.find('</sup>')
            title = title.replace('</sup>', '')
        
        return title


    # Change a date object like '2021-03-04 00:00:00' to '03 thru 04'
    # Excel mistakenly reads things like '3-4' as dates and we don't want that
    def date_to_num(input):
        input = input.split() #separate date from timestamp ['2021-03-04','00:00:00']
        input = input[0].split('-', 1) #separate year from month&day ['2021', '03-04']
        input = input[1].split('-') #separte month and day ['03, 04']

        #cast to and int and then back to a string to get rid of leading 0's
        if int(input[0]) < int(input[1]):
            return str(int(input[0])) + ' thru ' + str(int(input[1]))
        else:
            return str(int(input[1])) + ' thru ' + str(int(input[0]))


    # Harvesting Main Function
    def harvesting_main(file_name):
        update_progress(3)
        file_name = f"{file_name[:-5]}_Author_Split.xlsx"

        # Load Excel File
        old_path = str(file_name)
        old_book = load_workbook(old_path)  #read file
        old_sheet = old_book.active         #read sheet

        # Create a workbook and worksheet
        new_book = Workbook()               #create workbook
        new_sheet = new_book.active         #create worksheet

        #Read in column headers
        f = open('R:/storage/libarchive/a/zzz_Programs/Control Panel/Documentation/headers.txt', 'r')
        headers = f.read()
        f.close()

        headers = headers.replace(' ', '')
        headers = headers.split(',')

        # Add Column Headers (before authors)
        new_col = 1
        for header in headers:
            new_sheet.cell(1, new_col).value = header
            new_col += 1

        # Determine how many columns are in the sheet that is being read from
        old_max_col = len([c for c in old_sheet.iter_cols(min_row=1, max_row=1, values_only=True) if c[0] is not None])

        # Determine how many rows are in the sheet that is being read from
        old_max_row = len([c for c in old_sheet.iter_rows(min_col=1, max_col=1, values_only=True) if c[0] is not None])
        new_max_row = old_max_row

        # Starting author count
        last_cell = str(old_sheet.cell(1, old_max_col - 3).value)[6:]

        index = 0
        for letter in last_cell:
            if letter == '_':
                author_count = int(last_cell[:index])
                break
            index += 1

        # Adds the 7 headers for an author
        def add_author_headers(new_col, num): #takes in starting column and author number
            #add headers to worksheet
            new_sheet.cell(1,  new_col).value = f'author{num}_fname'
            new_sheet.cell(1,  new_col+1).value = f'author{num}_mname'
            new_sheet.cell(1,  new_col+2).value = f'author{num}_lname'
            new_sheet.cell(1,  new_col+3).value = f'author{num}_suffix'
            new_sheet.cell(1,  new_col+4).value = f'author{num}_email'
            new_sheet.cell(1,  new_col+5).value = f'author{num}_institution'
            new_sheet.cell(1,  new_col+6).value = f'author{num}_is_corporate'

        # Add authors
        for num in range(1, author_count + 1):      #creates column headers for author count
            add_author_headers(new_col, num)        #add column headers for another author
            new_col += 7                            #update the current empty column header

        # Determine how many columns are in the sheet that is being written to
        new_max_col = len([c for c in new_sheet.iter_cols(min_row=1, max_row=1, values_only=True) if c[0] is not None])

        def get_new_index(new_header):
            for new_col in range(1, new_max_col + 1):
                new_col_head = new_sheet.cell(1, new_col).value
                if new_header == new_col_head:
                    return new_col

        def get_old_index(old_header):
            for old_col in range(1, old_max_col + 1):
                old_col_head = old_sheet.cell(1, old_col).value
                if old_header == old_col_head:
                    return old_col

        # Copy Information
        # col is the column index where info is being copied to (new excel)
        # old_col is the column index where info is being copied from (old excel)
        def copy(new_header, old_header):
            try:
                new_col_index = get_new_index(new_header)
                old_col_index = get_old_index(old_header)
                new_sheet.cell(row, new_col_index).value = old_sheet.cell(row, old_col_index).value
            except:
                if row == 2:
                    warning_popup("Couldn't copy new '' + str(new_header) + '' column from old '' + str(old_header) + '' column.")

        # Fill Information
        # col is the column index where info is being filled (new excel)
        # fill_text is the text that is filled into the new excel
        def fill(new_header, fill_text):
            try:
                new_col_index = get_new_index(new_header)
                new_sheet.cell(row, new_col_index).value = fill_text
            except: 
                if row == 2:
                    warning_popup("Couldn't fill new '' + str(new_header) + '' column with '' + str(fill_text) + ''.")

        # Creates yellow background fill
        yellow = PatternFill(patternType = 'solid', fgColor = 'fffb00')

        # Add Highlight to Column Headers
        new_sheet.cell(1, get_new_index('open_access')).fill = yellow
        new_sheet.cell(1, get_new_index('url')).fill = yellow

        #Pathway count
        try:
            looking = True
            pathway_count = 0
            while(looking):
                if get_old_index(f'Pathway {pathway_count+1}: Version'):
                    #Found a set of pathway information
                    pathway_count += 1
                    looking = True
                else:
                    #Did not find anymore pathway information
                    looking = False
        except:
            warning_popup("Couldn't determine number of pathways. The 'rights', 'distribution_license', and 'embargo_date' columns will not transfer.")

        # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
        #          FILLS ROW INFORMATION (in order of reformatted excel)            #
        # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
        def row_information(row):
            # open_access
            if old_sheet.cell(row, get_old_index('Open Access')).value == 'OA':
                #Finds which specific OA it is
                if old_sheet.cell(row, get_old_index('Gold OA')).value == 'OA':
                    copy('open_access', 'Gold OA')
                elif old_sheet.cell(row, get_old_index('Hybrid Gold OA')).value != '-':
                    copy('open_access', 'Hybrid Gold OA')
                elif old_sheet.cell(row, get_old_index('Bronze OA')).value != '-':
                    copy('open_access', 'Bronze OA')
                elif old_sheet.cell(row, get_old_index('Green Final OA')).value != '-':
                    copy('open_access', 'Green Final OA')
                elif old_sheet.cell(row, get_old_index('Green Accepted OA')).value != '-':
                    copy('open_access', 'Green Accepted OA')
            else:
                #default OA
                copy('open_access', 'Open Access')

            # sherpa_romeo_uri
            # copy('sherpa_romeo_uri', 'Sherpa Romeo URI')

            # pathway_version
            # copy('pathway_version', 'Pathway 1: Version')

            # url
            copy('url', 'source_fulltext_url')

            # source_publication
            copy('source_publication', 'source_publication')

            # title
            copy('title', 'title')

            # doi
            copy('doi', 'doi')

            # source_fulltext_url
            copy('source_fulltext_url', 'source_fulltext_url')

            # additional_text_uri
            try:
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    version = old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value
                    #Check if published
                    if version == 'Published Version':
                        copy('additional_text_uri', 'source_fulltext_url')
                        fill('source_fulltext_url', '')
                        break
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'additional_text_uri' column.")
                
            # author_classification
            fill('author_classification', 'faculty')
                
            # total_author_count
            copy('total_author_count', 'total_author_count')
            
            # faculty_author_count
            copy('faculty_author_count', 'faculty_author_count')
            
            # institution_name
            fill('institution_name', 'Missouri University of Science and Technology')
            
            # authorized_name
            copy('authorized_name', 'authorized_name')
            
            # abstract
            copy('abstract', 'abstract')
            
            # department1
            copy('department1', 'department1')
            
            # department2
            copy('department2', 'department2')
            
            # department3
            copy('department3', 'department3')
            
            # department4
            copy('department4', 'department4')
            
            # Funding Number & Funding Sponsor
            try:
                f_num = old_sheet.cell(row, get_old_index('Funding Number')).value # Changing to grant?
                f_spon = old_sheet.cell(row, get_old_index('Funding Sponsor')).value # Changing to fundref?
                output = ''

                if f_spon != None:
                    if 'undefined' not in str(f_num):
                        output = str(f_spon) + ', Grant ' + str(f_num)
                    else:
                        output = str(f_spon)

                fill('comments', output)
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'Funding Number' and 'Funding Sponsor' columns.")

            # keywords
            copy('keywords', 'keywords')

            # isbn
            try:
                isbn = str(old_sheet.cell(row, get_old_index('isbn')).value) #get value

                if isbn != None:
                    #Get rid of spaces and brackets
                    isbn = isbn.replace('[', '')
                    isbn = isbn.replace(']', '')
                    isbn = isbn.replace(' ', '')

                    #make a list
                    isbn_list = isbn.split(',')

                    #validate number and format (must have 978 in the beginning)
                    index = 0
                    size = len(isbn_list)
                    while (index < size):
                        if isbn_list[index].find('978', 0, 3) == -1:
                            isbn_list.remove(isbn_list[index])
                            size -= 1
                        elif len(isbn_list[index]) != 13:
                            isbn_list.remove(isbn_list[index])
                            size -= 1
                        else:
                            isbn_list[index] = str(isbn_list[index][:3]) + '-' + str(isbn_list[index][3:12]) + '-' + str(isbn_list[index][12:])
                            index += 1

                    #fill the list with the numbers seperated by a semicolon if there are multiple
                    fill('isbn', ';'.join(isbn_list)) #output to cell
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'isbn' column.")

            # issn
            try:
                #initialization
                e_ISSN = old_sheet.cell(row, get_old_index('e-ISSN')).value
                issn = old_sheet.cell(row, get_old_index('issn')).value
                issn_values = []
                
                #add valid values
                if e_ISSN != None:
                    issn_values.append(e_ISSN)
                if issn != None:
                    issn_values.append(issn)

                #format
                for index in range(len(issn_values)):
                    issn_values[index] = str(issn_values[index]).zfill(8) #add missing leading 0s
                    parts = [issn_values[index][:4], issn_values[index][4:]] #[first 4 numbers, last 4 numbers]
                    issn_values[index] = '-'.join(parts) #format output ####-####

                fill('issn', '; '.join(issn_values))      #output to cell
            except TypeError:
                if row == 2:
                    warning_popup("Couldn't transfer 'issn' column.")

            # document_type
            copy('document_type', 'document_type')
            
            if 'conference' in str(new_sheet.cell(row, get_new_index('document_type')).value):
                fill('document_type', 'article_conference_proceedings')

            # document_version
            try:
                fill('document_version', 'Citation')
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    version = old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value
                    #Check if published
                    if version == 'Published Version':
                        fill('document_version', 'Final Version')
                        break
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'document_version' column.")
            
            # file_type
            fill('file_type', 'text')
            
            # language_iso
            fill('language_iso', 'English')
            
            # language2
            copy('language2', 'language2')
            
            # rights
            try:
                current_year = datetime.today().year
                rights = ''

                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    #Check if published
                    if old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value == 'Published Version':
                        if old_sheet.cell(row, get_old_index(f'Pathway {num}: Copyright Owners')).value and old_sheet.cell(row, get_old_index(f'Pathway {num}: Copyright Owners')).value != 'Publishers':
                            copyright_owners = old_sheet.cell(row, get_old_index(f'Pathway {num}: Copyright Owners')).value

                            if copyright_owners == 'Authors':
                                copyright_owners = 'The Authors'

                            rights = '© ' + str(current_year) + ' ' + str(copyright_owners) + ', All rights reserved.'
                            break
                        break
                    elif str(old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value) == 'None':
                        break

                # Rights has not been filled yet    
                if rights == '':
                    #Get initial value
                    publishers = str(old_sheet.cell(row, get_old_index('Publisher(s)')).value)

                    #Check if empty
                    if(publishers.count('None') > 0):
                        rights = '© ' + str(current_year) + ', All rights reserved.'
                    else:
                        #Initialize list
                        publisher_list = [publishers]

                        #Split if multiple publishers
                        if(publishers.count(';') > 0):
                            publisher_list = publishers.split('; ')

                        #Get rid of unwanted info
                        for index in range(len(publisher_list)):
                            publisher_list[index] = publisher_list[index].split(' [')[0]

                        #Combine if multiple
                        publishers = '; '.join([str(elem) for elem in publisher_list])

                        rights = '© ' + str(current_year) + ' ' + str(publishers) + ', All rights reserved.'

                fill('rights', rights)
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'rights' column.")

            # distribution_license 
            try:
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    #Check if published
                    if old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value == 'Published Version':
                        if old_sheet.cell(row, get_old_index(f'Pathway {num}: License')).value:
                            #Get info from old sheet
                            license_tag = old_sheet.cell(row, get_old_index(f'Pathway {num}: License')).value[3:].lower().split()
                            if len(license_tag) == 1:
                                license_tag.append('4')
                            license_link = 'http://creativecommons.org/licenses/' + str(license_tag[0]) + '/' + str(license_tag[1][0]) + '.0/'
                            fill('distribution_license', license_link)
                        break
            except:
                warning_popup("Couldn't transfer 'distribution_license' column.")
            
            # publication_date
            copy('publication_date', 'publication_date')
            
            try:
                if new_sheet.cell(row, get_new_index('publication_date')).value:
                    date = str(new_sheet.cell(row, get_new_index('publication_date')).value)
                    if len(date) > 10:
                        date = date.split() #gets rid of time in the format
                        fill('publication_date', date[0])
            except:
                warning_popup("Couldn't transfer 'publication_date' column.")

            # custom_publication_date
            try:
                date = str(new_sheet.cell(row, get_new_index('publication_date')).value)

                if date != 'None':
                    date = date.split()
                    date = date[0] #gets rid of time in the format

                    # Get parts from date
                    date = date.split('-') #WORKS WHEN FORMAT IS '2020-02-01'

                    # Get month in letters
                    x = datetime(int(date[0]), int(date[1]), int(date[2]))
                    month = x.strftime('%b')

                    # Output
                    fill('custom_publication_date', date[2] + ' ' + month + ' ' + date[0])
            except:
                warning_popup("Unusual formatting in 'publication_date' column")

            # publisher
            try:
                #Get initial value
                publishers = str(old_sheet.cell(row, get_old_index('Publisher(s)')).value)
                
                #Check if empty
                if(publishers.count('None') > 0):
                    publishers = ''
                else:
                    #Initialize list
                    publisher_list = [publishers]
                  
                    #Split if multiple publishers
                    if(publishers.count(';') > 0):
                        publisher_list = publishers.split('; ')
                       
                    #Get rid of unwanted info
                    for index in range(len(publisher_list)):
                        publisher_list[index] = publisher_list[index].split(' [')[0]
                       
                    #Combine if multiple
                    publishers = '; '.join([str(elem) for elem in publisher_list])
                    
                # Output
                fill('publisher', publishers)
            except:
                warning_popup("Could not transfer 'Publisher(s)' column")
            
            # volnum
            copy('volnum', 'volnum')
            
            # issnum
            copy('issnum', 'issnum')
            
            # articlenum
            copy('articlenum', 'articlenum')
            
            # fpage
            copy('fpage', 'fpage')
            
            # lpage
            copy('lpage', 'lpage')
            
            # pubmedid
            copy('pubmedid', 'pubmedid')
            
            # embargo_date
            try:
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    #Check if published
                    if old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value == 'Published Version':
                        #Get info from old sheet
                        embargo_offset = old_sheet.cell(row, get_old_index(f'Pathway {num}: Embargo')).value.split()
                        pub_date = old_sheet.cell(row, get_old_index('publication_date')).value

                        #Embargo date is the same as publisher date
                        if embargo_offset[0].count('No') > 0:
                            fill('embargo_date', str(pub_date))
                                                    
                        #Embargo date is offset from the publisher date by an amount of months
                        elif embargo_offset[1].count('Months') > 0:
                            #Add embargo offset to the published date to get the embargo date
                            embargo_date = pub_date + relativedelta(months = int(embargo_offset[0]))
                            fill('embargo_date', str(embargo_date))
                        break
            except TypeError:
                if row == 2:
                    warning_popup("Couldn't transfer 'embargo_date' column.")

            # primary_document_attached
            try:
                fill('primary_document_attached', 'no')
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    version = old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value
                    #Check if published
                    if version == 'Published Version':
                        fill('primary_document_attached', 'yes')
                        break
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'primary_document_attached' column.")

            # copy author information
            try:
                a_count = old_sheet.cell(row, get_old_index('total_author_count')).value
                for num in range(1, a_count + 1):
                    copy(f'author{num}_fname', f'author{num}_fname')
                    copy(f'author{num}_mname', f'author{num}_mname')
                    copy(f'author{num}_lname', f'author{num}_lname')
                    copy(f'author{num}_suffix', f'author{num}_suffix')
                    copy(f'author{num}_email', f'author{num}_email')
                    copy(f'author{num}_institution', f'author{num}_institution')
                    copy(f'author{num}_is_corporate', f'author{num}_is_corporate')
            except:
                warning_popup('An error occured while copying author information. Excel may be incomplete.')

        # Transfer row information for the whole sheet
        update_progress(4) # Filling in ' + str(old_max_row) + ' rows

        for row in range(2, old_max_row + 1):
            row_information(row)

        # Fix All Uppercase Titles
        update_progress(5) # Fixing all uppercase titles

        try:
            for row in range(2, new_max_row):
                # Get the title
                title = new_sheet.cell(row, get_new_index('title')).value
                # Split the title on the spaces
                title = title.split()
                # Capitalize each first letter
                cap_title = str()
                for index in range(len(title)):
                    cap_title = cap_title + title[index][0].capitalize() + title[index][1:] + ' '
                #Replace uncapitalized Title
                fill('title', cap_title[:-1])
        except:
            warning_popup('An error occured while fixing uppercase titles. Excel may be incomplete.')

        # Fix Subscripts and
        update_progress(6) # Fixing subscripts and superscripts

        try:
            for row in range(2, new_max_row):
                fill('title', replace_sub_sup(new_sheet.cell(row, get_new_index('title')).value))
        except:
            warning_popup('An error occured while fixing subscripts and superscripts. Excel may be incomplete.')

        # Switching All Dates to Numbers
        update_progress(7) # Switching all dates to numbers

        try:
            for row in range(2, new_max_row):
                # Fix issnum column
                col_value = str(new_sheet.cell(row, get_new_index('issnum')).value)
                if col_value:                                                   #if a value exists
                    if '00:00:00' in col_value:                                 #change date to number
                        fill('issnum', date_to_num(str(col_value)))
                    elif '-' in col_value:                                      #get rid of dash
                        fill('issnum', str(col_value).replace('-', ' thru '))

                    if col_value.isalpha(): #get rid of words
                        fill('issnum', '')

                # Fix volnum column
                col_value = str(new_sheet.cell(row, get_new_index('volnum')).value)
                if col_value:                                                   #if a value exists
                    if '00:00:00' in col_value:                                 #change date to number
                        fill('volnum', date_to_num(str(col_value)))
                    elif '-' in col_value:                                      #get rid of dash
                        fill('volnum', str(col_value).replace('-', ' thru '))

                    if 'a' in col_value or 'e' in col_value or 'i' in col_value or 'o' in col_value or 'u' in col_value or 'y' in col_value: #get rid of words
                        fill('volnum', '')
        except:
            warning_popup('An error occured while fixing date formats. Excel may be incomplete.')

        # Format Columns
        update_progress(8) # Adjusting column width

        try:
            for new_col in range(1, new_max_col + 1):
                max_length = 0
                for new_row in range(1, new_max_row):                               #go through columns and rows
                    if len(str(new_sheet.cell(new_row, new_col).value)) > max_length:
                        max_length = len(str(new_sheet.cell(new_row, new_col).value))   #find the longest cell
                        if max_length >= 100:                                       #cell cannont be longer than 100
                            new_sheet.column_dimensions[get_column_letter(new_col)].width = 100
                            break

                # Adjust the cell length
                if max_length < 100:
                    new_sheet.column_dimensions[get_column_letter(new_col)].width = max_length * 1.13
        except:
            warning_popup('An error occured while adjusting column width. Some column widths may not have changed.')

        # Save excel
        update_progress(9) # Saving

        try:
            new_path = str(file_name)[:len(file_name) - 18] + '_Completed.xlsx'
            new_book.save(new_path)

            update_progress(10)
        except:
            new_path = str(file_name)[:len(file_name) - 18] + '_Completed.xlsx'
            saved_name = new_path.split('/')
            error_popup(str(saved_name[-1]) + "' could not be saved.")



    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
    #                                      Main                                         #
    # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

    # Customize window
    Both.title('Excel Processing Program')
    Both.configure(bg = '#003B49')
    Both.iconbitmap('R:/storage/libarchive/a/zzz_Programs/Control Panel/CODE/S&T_Logo.ico')

    # Place in window
    top_frame.pack(padx = 20, pady = 10)
    middle_frame.pack(padx = 20, pady = 0, fill = BOTH, expand = 1)
    bottom_frame.pack(padx = 20, pady = 10)

    # Place in top_frame
    help_button.pack(side = RIGHT, padx = 5)
    select_button.pack(side = LEFT, padx = 5)

    # Place in middle_frame
    scroll_bar.pack(side = RIGHT, fill = Y)
    label_canvas.pack(side = LEFT, fill = BOTH, expand = 1)

    # Place in bottom_frame
    progress.grid(row = 1, column = 0, padx = 5, pady = 5)
    start_button.grid(row = 1, column = 1, padx = 5, pady = 5)

    # Configure
    label_canvas.configure(yscrollcommand = scroll_bar.set)
    label_canvas.bind('<Configure>', lambda e: label_canvas.configure(scrollregion = label_canvas.bbox("all")))
    label_canvas.bind_all('<MouseWheel>', lambda e: label_canvas.yview('scroll', (-1*e.delta), "units"))

    # Place in Canvas
    label_canvas.create_window((0, 0), window = canvas_frame, anchor = 'nw')

    # Place in canvas_frame
    file_label.pack(fill = BOTH)

    # Keeps window open until closed
    Both.mainloop()