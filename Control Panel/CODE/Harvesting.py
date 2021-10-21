# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           IMPORT                                                                #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
import datetime
from dateutil.relativedelta import relativedelta
import sqlite3
import time
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, PatternFill
from openpyxl.styles.colors import Color
from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import os



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           CHANGE DATE TO NUM                                                      #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def date_to_num(input):
    input = input.split() #separate date from timestamp ["2021-03-04","00:00:00"]
    input = input[0].split("-", 1) #separate year from month&day ["2021", "03-04"]
    input = input[1].split("-") #separte month and day ["03, 04"]

    #cast to and int and then back to a string to get rid of leading 0's
    if int(input[0]) < int(input[1]):
        return str(int(input[0])) + " thru " + str(int(input[1]))
    else:
        return str(int(input[1])) + " thru " + str(int(input[0]))



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           FIX SUBSCRIPT AND SUPERSCRIPT                                           #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def replace_sub_sup(title):
    while('</sub>' in title):
        start = title.find('<sub>')
        end = title.find('</sub>')
        title = title.replace('<sub>', "")
        while start != end:
            title = str(title[0:start]) + str(char_to_sub(title[start])) + str(title[start + 1:])
            start+=1
            end = title.find('</sub>')
        title = title.replace('</sub>', "")

    while('</sup>' in title):
        start = title.find('<sup>')
        end = title.find('</sup>')
        title = title.replace('<sup>', "")
        while start != end:
            title = str(title[0:start]) + str(char_to_sup(title[start])) + str(title[start + 1:])
            start+=1
            end = title.find('</sup>')
        title = title.replace('</sup>', "")
    
    return title



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           CHARACTER TO SUBSCRIPT                                                  #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def char_to_sub(argument):
    output = ''
    switcher = {
        '0': '\u2080',
        '1': '\u2081',
        '2': '\u2082',
        '3': '\u2083',
        '4': '\u2084',
        '5': '\u2085',
        '6': '\u2086',
        '7': '\u2087',
        '8': '\u2088',
        '9': '\u2089',
        '+': '\u208A',
        '-': '\u208B',
        '=': '\u208C',
        '(': '\u208D',
        ')': '\u208E',
        'a': '\u2090',
        'e': '\u2091',
        'o': '\u2092',
        'x': '\u2093',
        'h': '\u2095',
        'k': '\u2096',
        'l': '\u2097',
        'm': '\u2098',
        'n': '\u2099',
        'p': '\u209A',
        's': '\u209B',
        't': '\u209C',
    }
    output = switcher.get(argument, '?')
    if '?' in output:
        SUB = str.maketrans("aehijklmnoprstuvxy", "ₐₑₕᵢⱼₖₗₘₙₒₚᵣₛₜᵤᵥₓᵧ")
        output = argument.translate(SUB)
    return output



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           CHARACTER TO SUPERSCRIPT                                                #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def char_to_sup(argument):
    output = ''
    switcher = {
        '0': '\u2070',
        '1': '\u00B9',
        '2': '\u00B2',
        '3': '\u00B3',
        '4': '\u2074',
        '5': '\u2075',
        '6': '\u2076',
        '7': '\u2077',
        '8': '\u2078',
        '9': '\u2079',
        'i': '\u2071',
        '+': '\u207A',
        '-': '\u207B',
        '=': '\u207C',
        '(': '\u207D',
        ')': '\u207E',
        'n': '\u207F',
    }
    output = switcher.get(argument, '?')
    if '?' in output:
        SUP = str.maketrans("ABDEGHIJKLMNOPRTUVWabcdefghijklmnoprstuvwxyz", "ᴬᴮᴰᴱᴳᴴᴵᴶᴷᴸᴹᴺᴼᴾᴿᵀᵁⱽᵂᵃᵇᶜᵈᵉᶠᵍʰⁱʲᵏˡᵐⁿᵒᵖʳˢᵗᵘᵛʷˣʸᶻ")
        output = argument.translate(SUP)
    return output



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           DISPLAY WINDOW                                                          #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def open_harvesting():
    # Create window
    harvesting = Toplevel()
    harvesting.title("Harvesting Program")
    harvesting.configure(bg = '#003B49')
    harvesting.iconbitmap('R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Control Panel/CODE/S&T_Logo.ico')

    # Center the window on the screen
    window_w = 400 # window width
    window_h = 290 # window height

    screen_w = harvesting.winfo_screenwidth() # screen width
    screen_h = harvesting.winfo_screenheight() # screen height

    x_coor = (screen_w/2) - (window_w/2) #middle of screen x coordinate
    y_coor = (screen_h/2) #lower half of screen y coordinate

    harvesting.geometry("%dx%d+%d+%d" % (window_w, window_h, x_coor, y_coor)) # place in middle

    #Create Progress Bar
    progress = ttk.Progressbar(harvesting, orient = HORIZONTAL, length = 370, mode = 'determinate')

    # Create a frame
    frame = LabelFrame(harvesting, text = "Select an Input File", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')

    # Create a label
    task = Label(harvesting, text = "Waiting for a file", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')
    file_label = Label(harvesting, text = "File: N/A", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')

    # Open Help Function
    def open_help():
        # Open word document
        try:
            os.startfile("**/*/Control Panel/Documentation/Harvesting Help.docx")
        except:
            error_popup("Could not find a help file to open.")


    # Create a button
    help_button = Button(harvesting, text = "Help", command = lambda : open_help(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    exit_button = Button(harvesting, text = "Exit Harvesting", command = harvesting.destroy, bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")

    # Place in window
    frame.pack(padx = 20, pady = 10)
    file_label.pack(padx = 20, pady = (0, 20))
    progress.pack(padx = 5, pady = 5)
    task.pack(padx = 20)
    help_button.pack(padx = (331, 10), pady = 0)
    exit_button.pack(padx = (250, 10), pady = (5, 0))

    # Error Message Popup
    def error_popup(error_message):
        messagebox.showerror("Error", error_message)

    # Warning Message Popup
    def warning_popup(warning_message):
        messagebox.showwarning("Warning", warning_message)

    # Information Message Popup
    def info_popup(info_message):
        messagebox.showinfo("Information", info_message)

    # Update Progress Bar
    def update_progress(p, t):
        # Update bar value
        progress['value'] = (p/9)*100

        # Update bar label
        task.config(text = t)
        task.pack()

        # Refresh window (very important line!)
        harvesting.update_idletasks()
        time.sleep(0.2)

    # Function Click
    def browse():
        # Reset progress bar
        update_progress(0, "Waiting for a file")

        # Open file
        harvesting.filename = filedialog.askopenfilename(initialdir = "R:/storage/libarchive/b/1. Processing/8. Other Projects/Excel Files", title = "Select Input", filetypes = (("Excel Workbook", "*.xlsx"),))
        if harvesting.filename == "":
            warning_popup("No file selected.")
            file_label.config(text = "File: N/A")
            del harvesting.filename
        else:
            # Get file name
            name = harvesting.filename
            for i in range(len(name)):
                if "/" in name[-(i)]:
                    name = "File: " + name[-(i-1):]
                    break

            # Update Folder Name Label
            file_label.config(text = name)

    # Main Function
    def main(file_name):
        update_progress(1, "Harvesting files...")

        # Load Excel File
        old_path = str(file_name)
        old_book = load_workbook(old_path)  #read file
        old_sheet = old_book.active         #read sheet

        # Create a workbook and worksheet
        new_book = Workbook()               #create workbook
        new_sheet = new_book.active         #create worksheet

        # Column Headers
        headers =   ["open_access", "sherpa_romeo_uri", "pathway_version", "url", "title", "title_alternative", "doi", "source_fulltext_url",
                     "additional_text_uri", "author_classification", "total_author_count", "faculty_author_count", 
                     "institution_name", "authorized_name", "abstract", "meeting_name", "department1", "department2", 
                     "department3", "department4", "centers_labs", "centers_labs2", "centers_labs3", "centers_labs4", 
                     "pub_status", "comments", "keywords", "geographic_coverage", "time_period", "isbn", "issn", 
                     "document_type", "document_version", "file_type", "language_iso", "language2", "rights", 
                     "distribution_license", "publication_date", "custom_publication_date", "publisher",
                     "publisher_place", "source_publication", "volnum", "issnum", "articlenum", "fpage", 
                     "lpage", "pubmedid", "disciplines", "embargo_date", "date_uploaded", 
                     "primary_document_attached", "copyright"]

        # Add Column Headers (before authors)
        new_col = 1
        for header in headers:
            new_sheet.cell(1, new_col).value = header
            new_col += 1

        # Determine how many columns are in the sheet that is being read from
        old_max_col = len([c for c in old_sheet.iter_cols(min_row=1, max_row=1, values_only=True) if c[0] is not None])

        # Determine how many rows are in the sheet that is being read from
        old_max_row = len([c for c in old_sheet.iter_rows(min_col=1, max_col=1, values_only=True) if c[0] is not None])
        new_max_row = old_max_row

        # Starting author count
        last_cell = str(old_sheet.cell(1, old_max_col - 3).value)[6:]

        index = 0
        for letter in last_cell:
            if letter == '_':
                author_count = int(last_cell[:index])
                break
            index += 1

        # Adds the 7 headers for an author
        def add_author_headers(new_col, num): #takes in starting column and author number
            #add headers to worksheet
            new_sheet.cell(1,  new_col).value = f'author{num}_fname'
            new_sheet.cell(1,  new_col+1).value = f'author{num}_mname'
            new_sheet.cell(1,  new_col+2).value = f'author{num}_lname'
            new_sheet.cell(1,  new_col+3).value = f'author{num}_suffix'
            new_sheet.cell(1,  new_col+4).value = f'author{num}_email'
            new_sheet.cell(1,  new_col+5).value = f'author{num}_institution'
            new_sheet.cell(1,  new_col+6).value = f'author{num}_is_corporate'

        # Add authors
        for num in range(1, author_count + 1):      #creates column headers for author count
            add_author_headers(new_col, num)        #add column headers for another author
            new_col += 7                            #update the current empty column header

        # Determine how many columns are in the sheet that is being written to
        new_max_col = len([c for c in new_sheet.iter_cols(min_row=1, max_row=1, values_only=True) if c[0] is not None])

        def get_new_index(new_header):
            for new_col in range(1, new_max_col + 1):
                new_col_head = new_sheet.cell(1, new_col).value
                if new_header == new_col_head:
                    return new_col

        def get_old_index(old_header):
            for old_col in range(1, old_max_col + 1):
                old_col_head = old_sheet.cell(1, old_col).value
                if old_header == old_col_head:
                    return old_col

        # Copy Information
        # col is the column index where info is being copied to (new excel)
        # old_col is the column index where info is being copied from (old excel)
        def copy(new_header, old_header):
            try:
                new_col_index = get_new_index(new_header)
                old_col_index = get_old_index(old_header)
                new_sheet.cell(row, new_col_index).value = old_sheet.cell(row, old_col_index).value
            except:
                if row == 2:
                    warning_popup("Couldn't copy new '" + str(new_header) + "' column from old '" + str(old_header) + "' column.")

        # Fill Information
        # col is the column index where info is being filled (new excel)
        # fill_text is the text that is filled into the new excel
        def fill(new_header, fill_text):
            try:
                new_col_index = get_new_index(new_header)
                new_sheet.cell(row, new_col_index).value = fill_text
            except: 
                if row == 2:
                    warning_popup("Couldn't fill new '" + str(new_header) + "' column with '" + str(fill_text) + "'.")

        # Creates yellow background fill
        yellow = PatternFill(patternType = 'solid', fgColor = 'fffb00')

        # Add Highlight to Column Headers
        new_sheet.cell(1, get_new_index('open_access')).fill = yellow
        new_sheet.cell(1, get_new_index('url')).fill = yellow
        new_sheet.cell(1, get_new_index('copyright')).fill = yellow

        #Pathway count
        try:
            looking = True
            pathway_count = 0
            while(looking):
                if get_old_index(f'Pathway {pathway_count+1}: Version'):
                    #Found a set of pathway information
                    pathway_count += 1
                    looking = True
                else:
                    #Did not find anymore pathway information
                    looking = False
        except:
            warning_popup("Could not determine number of pathways. The 'rights', 'distribution_license', and 'embargo_date' columns will not transfer.")

        # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
        #          FILLS ROW INFORMATION (in order of reformatted excel)            #
        # * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
        def row_information(row):
            # open_access
            if old_sheet.cell(row, get_old_index('Open Access')).value == "OA":
                #Finds which specific OA it is
                if old_sheet.cell(row, get_old_index('Gold OA')).value == "OA":
                    copy('open_access', 'Gold OA')
                elif old_sheet.cell(row, get_old_index('Hybrid Gold OA')).value != "-":
                    copy('open_access', 'Hybrid Gold OA')
                elif old_sheet.cell(row, get_old_index('Bronze OA')).value != "-":
                    copy('open_access', 'Bronze OA')
                elif old_sheet.cell(row, get_old_index('Green Final OA')).value != "-":
                    copy('open_access', 'Green Final OA')
                elif old_sheet.cell(row, get_old_index('Green Accepted OA')).value != "-":
                    copy('open_access', 'Green Accepted OA')
            else:
                #default OA
                copy('open_access', 'Open Access')


            # sherpa_romeo_uri
            # copy('sherpa_romeo_uri', 'Sherpa Romeo URI')

            # pathway_version
            # copy('pathway_version', 'Pathway 1: Version')

            # url
            copy('url', 'source_fulltext_url')

            # title
            copy('title', 'title')

            # doi
            copy('doi', 'doi')

            # source_fulltext_url
            copy('source_fulltext_url', 'source_fulltext_url')
            
            # author_classification
            fill('author_classification', 'faculty')
                
            # total_author_count
            copy('total_author_count', 'total_author_count')
            
            # faculty_author_count
            copy('faculty_author_count', 'faculty_author_count')
            
            # institution_name
            fill('institution_name', 'Missouri University of Science and Technology')
            
            # authorized_name
            copy('authorized_name', 'authorized_name')
            
            # abstract
            copy('abstract', 'abstract')
            
            # department1
            copy('department1', 'department1')
            
            # department2
            copy('department2', 'department2')
            
            # department3
            copy('department3', 'department3')
            
            # department4
            copy('department4', 'department4')
            
            # Funding Number & Funding Sponsor
            try:
                f_num = old_sheet.cell(row, get_old_index('Funding Number')).value # Changing to grant?
                f_spon = old_sheet.cell(row, get_old_index('Funding Sponsor')).value # Changing to fundref?
                output = ''

                if f_spon != None:
                    if 'undefined' not in str(f_num):
                        output = str(f_spon) + ', Grant ' + str(f_num)
                    else:
                        output = str(f_spon)

                fill('comments', output)
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'Funding Number' and 'Funding Sponsor' columns.")

            # keywords
            copy('keywords', 'keywords')

            # isbn
            try:
                isbn = str(old_sheet.cell(row, get_old_index('isbn')).value) #get value

                if isbn != None:
                    #Get rid of spaces and brackets
                    isbn = isbn.replace("[", "")
                    isbn = isbn.replace("]", "")
                    isbn = isbn.replace(" ", "")

                    #make a list
                    isbn_list = isbn.split(",")

                    #validate number and format (must have 978 in the beginning)
                    index = 0
                    size = len(isbn_list)
                    while (index < size):
                        if isbn_list[index].find("978", 0, 3) == -1:
                            isbn_list.remove(isbn_list[index])
                            size -= 1
                        elif len(isbn_list[index]) != 13:
                            isbn_list.remove(isbn_list[index])
                            size -= 1
                        else:
                            isbn_list[index] = str(isbn_list[index][:3]) + '-' + str(isbn_list[index][3:12]) + '-' + str(isbn_list[index][12:])
                            index += 1

                    #fill the list with the numbers seperated by a semicolon if there are multiple
                    fill('isbn', ";".join(isbn_list)) #output to cell
            except:
                if row == 2:
                    warning_popup("Couldn't transfer 'isbn' column.")

            # issn
            try:
                e_ISSN = old_sheet.cell(row, get_old_index('e-ISSN')).value
                issn = old_sheet.cell(row, get_old_index('issn')).value
                issn_values = []
                
                if e_ISSN != None:
                    issn_values.append(e_ISSN)
                if issn != None:
                    issn_values.append(issn)

                for index in range(len(issn_values)):
                    issn_values[index] = str(issn_values[index]).zfill(8) #add missing leading 0s
                    parts = [issn_values[index][:4], issn_values[index][4:]] #[first 4 numbers, last 4 numbers]
                    issn_values[index] = "-".join(parts) #format output ####-####

                fill('issn', "; ".join(issn_values))      #output to cell
            except TypeError:
                if row == 2:
                    warning_popup("Couldn't transfer 'issn' column.")

            # document_type
            copy('document_type', 'document_type')
            
            if 'conference' in str(new_sheet.cell(row, get_new_index('document_type')).value):
                fill('document_type', 'article_conference_proceedings')

            # document_version
            fill('document_version', 'Citation')
            
            # file_type
            fill('file_type', 'text')
            
            # language_iso
            fill('language_iso', 'English')
            
            # language2
            copy('language2', 'language2')
            
            # rights
            try:
                rights = '© 2021 , All rights reserved.'
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    #Check if published
                    if old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value == "Published Version":
                        if old_sheet.cell(row, get_old_index(f'Pathway {num}: Copyright Owners')).value:
                            copyright_oweners = old_sheet.cell(row, get_old_index(f'Pathway {num}: Copyright Owners')).value
                            rights = '© 2021 ' + str(copyright_oweners) + ', All rights reserved.'
                        break
                fill('rights', rights)
            except:
                if row == 2:
                    warning_popup("Could transfer 'rights' column.")

            # distribution_license 
            try:
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    #Check if published
                    if old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value == "Published Version":
                        if old_sheet.cell(row, get_old_index(f'Pathway {num}: License')).value:
                            #Get info from old sheet
                            license_tag = old_sheet.cell(row, get_old_index(f'Pathway {num}: License')).value[3:].lower().split()
                            if len(license_tag) == 1:
                                license_tag.append("4")
                            license_link = "http://creativecommons.org/licenses/" + str(license_tag[0]) + "/" + str(license_tag[1][0]) + ".0/"
                            fill('distribution_license', license_link)
                        break
            except:
                warning_popup("Could transfer 'distribution_license' column.")
            
            # publication_date
            copy('publication_date', 'publication_date')
            
            try:
                if new_sheet.cell(row, get_new_index('publication_date')).value:
                    date = str(new_sheet.cell(row, get_new_index('publication_date')).value)
                    if len(date) > 10:
                        date = date[:len(date) - 9] #gets rid of time in the format
                        fill('publication_date', date)
            except:
                print("")

            # custom_publication_date
            try:
                if new_sheet.cell(row, get_new_index('publication_date')).value:
                    date = str(new_sheet.cell(row, get_new_index('publication_date')).value)
                    date = date[:10] #gets rid of time in the format

                    #WORKS WHEN FORMAT IS "2020-02-01 00:00:00"
                    # Get parts from date
                    year = date[:4]
                    month = date[5:7]
                    day = date[8:]

                    # Get month in letters
                    x = datetime.datetime(int(year), int(month), int(day))
                    month = x.strftime("%b")

                    # Format output
                    output = day + ' ' + month + ' ' + year

                    # Output
                    fill('custom_publication_date', output)
            except:
                print("Unusual formatting in 'publication_date' column")

            # publisher
            copy('publisher', 'publisher')
            
            # volnum
            copy('volnum', 'volnum')
            
            # issnum
            copy('issnum', 'issnum')
            
            # articlenum
            copy('articlenum', 'articlenum')
            
            # fpage
            copy('fpage', 'fpage')
            
            # lpage
            copy('lpage', 'lpage')
            
            # pubmedid
            copy('pubmedid', 'pubmedid')
            
            # embargo_date
            try:
                #Loop through pathways
                for num in range(1, pathway_count + 1):
                    #Check if published
                    if old_sheet.cell(row, get_old_index(f'Pathway {num}: Version')).value == "Published Version":
                        #Get info from old sheet
                        embargo_offset = old_sheet.cell(row, get_old_index(f'Pathway {num}: Embargo')).value.split()
                        pub_date = old_sheet.cell(row, get_old_index('publication_date')).value
                        
                        #Embargo date is the same as publisher date
                        if embargo_offset[0] == "No":
                            fill('embargo_date', str(pub_date))
                        
                        #Embargo date is offset from the publisher date by an amount of months
                        elif embargo_offset[1] == "Months":
                            #Add embargo offset to the published date to get the embargo date
                            pub_date = pub_date.split("-")
                            embargo_date = datetime.date(int(pub_date[0]), int(pub_date[1]), int(pub_date[2])) + relativedelta(months = int(embargo_offset[0]))
                            fill('embargo_date', str(embargo_date))
                        break
            except TypeError:
                if row == 2:
                    warning_popup("Couldn't transfer 'embargo_date' column.")

            # primary_document_attached
            fill('primary_document_attached', 'no')

            # copy author information
            try:
                a_count = old_sheet.cell(row, get_old_index('total_author_count')).value
                for num in range(1, a_count + 1):
                    copy(f'author{num}_fname', f'author{num}_fname')
                    copy(f'author{num}_mname', f'author{num}_mname')
                    copy(f'author{num}_lname', f'author{num}_lname')
                    copy(f'author{num}_suffix', f'author{num}_suffix')
                    copy(f'author{num}_email', f'author{num}_email')
                    copy(f'author{num}_institution', f'author{num}_institution')
                    copy(f'author{num}_is_corporate', f'author{num}_is_corporate')
            except:
                warning_popup("An error occured while copying author information. Excel may be incomplete.")

        # Transfer row information for the whole sheet
        update_progress(2, "Harvesting files...") # Filling in " + str(old_max_row) + " rows

        for row in range(2, old_max_row + 1):
            row_information(row)

        # Fix All Uppercase Titles
        update_progress(3, "Harvesting files...") # Fixing all uppercase titles

        try:
            for row in range(2, new_max_row):
                title = new_sheet.cell(row, get_new_index('title')).value
                fill('title', title.title())
        except:
            warning_popup("An error occured while fixing uppercase titles. Excel may be incomplete.")

        # Fix Subscripts and
        update_progress(4, "Harvesting files...") # Fixing subscripts and superscripts

        try:
            for row in range(2, new_max_row):
                fill('title', replace_sub_sup(new_sheet.cell(row, get_new_index('title')).value))
        except:
            warning_popup("An error occured while fixing subscripts and superscripts. Excel may be incomplete.")

        # Switching All Dates to Numbers
        update_progress(5, "Harvesting files...") # Switching all dates to numbers

        try:
            for row in range(2, new_max_row):
                # Fix issnum column
                col_value = str(new_sheet.cell(row, get_new_index('issnum')).value)
                if col_value:                                                   #if a value exists
                    if '00:00:00' in col_value:                                 #change date to number
                        fill('issnum', date_to_num(str(col_value)))
                    elif '-' in col_value:                                      #get rid of dash
                        fill('issnum', str(col_value).replace("-", " thru "))

                    if col_value.isalpha(): #get rid of words
                        fill('issnum', '')

                # Fix volnum column
                col_value = str(new_sheet.cell(row, get_new_index('volnum')).value)
                if col_value:                                                   #if a value exists
                    if '00:00:00' in col_value:                                 #change date to number
                        fill('volnum', date_to_num(str(col_value)))
                    elif '-' in col_value:                                      #get rid of dash
                        fill('volnum', str(col_value).replace("-", " thru "))

                    if 'a' in col_value or 'e' in col_value or 'i' in col_value or 'o' in col_value or 'u' in col_value or 'y' in col_value: #get rid of words
                        fill('volnum', '')
        except:
            warning_popup("An error occured while fixing date formats. Excel may be incomplete.")

        # Format Columns
        update_progress(8, "Harvesting files...") # Adjusting column width

        try:
            for new_col in range(1, new_max_col + 1):
                max_length = 0
                for new_row in range(1, new_max_row):                               #go through columns and rows
                    if len(str(new_sheet.cell(new_row, new_col).value)) > max_length:
                        max_length = len(str(new_sheet.cell(new_row, new_col).value))   #find the longest cell
                        if max_length >= 100:                                       #cell cannont be longer than 100
                            new_sheet.column_dimensions[get_column_letter(new_col)].width = 100
                            break

                # Adjust the cell length
                if max_length < 100:
                    new_sheet.column_dimensions[get_column_letter(new_col)].width = max_length * 1.13
        except:
            warning_popup("An error occured while adjusting column width. Some column widths may not have changed.")

        # Save excel
        update_progress(9, "Harvesting files...") # Saving

        try:
            new_path = str(file_name)[:len(file_name) - 18] + '_Completed.xlsx'
            new_book.save(new_path)

            saved_name = new_path.split('/')
            info_popup("'" + str(saved_name[-1]) + "' has been saved.")

            update_progress(10, "Excel Created")
        except:
            error_popup("An error occured while saving the excel. The excel may not have been saved.")

    # Start Button Function
    def start():
        # Run program and update progress bar
        #try:
        main(harvesting.filename)
        #except PermissionError:
        #    error_popup("Could not save processed excel. The excel file that needs to be saved over is currently open. Close the excel file and hit the 'start' button again.")
        #except:
        #    error_popup("There was an unknown error, the file could not be processed.")

    # Create a button
    browse_button = Button(frame, text = "Browse", command = lambda : browse(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    start_button = Button(frame, text = "Start", command = lambda : start(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")

    # Place in frame
    browse_button.grid(row = 0, column = 0, padx = (15, 0), pady = 15)
    start_button.grid(row = 0, column = 1, padx = 15, pady = 15)
