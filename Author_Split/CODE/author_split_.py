import os
import ftfy
import openpyxl as xl
import pickle
import sys
_global = sys.modules[__name__]

rdsheet = None
author_column = ''
excelName = ''
authority_control_lookup_path = 'R:/storage/libarchive/a/Student Processing/zzz_Program_Dependencies/Authority_Control_Lookups.xlsx'
Authority_Lookup_Dict = pickle.load(open('Authority_Control_Dict_lowercase.pickle','rb'))
authority_control_lookup = None
authorDict = {}
rb = None

special_char = pickle.load(open('special_char.pickle','rb'))
extra_special_char = pickle.load(open('extra_special_char.pickle','rb'))



# Converts a character to it's index in alphabet where A = 0 and Z = 25
def alpha(letter):
    return ord(letter)-65

# Give excel location such as 'A5' and returns a (row, column) tuple 'A5' -> (5,0)
def xl_to_num(loc):
    loc = loc.split(':')
    if len(loc) == 1:
        col = loc[0]
        if len(col) == 1:
            return alpha(col)+1
        else:
            return (alpha(col[0])+1)*26+alpha(col[1])+1
    else:
        col,row = loc
        row = int(row)
        if len(col) == 1:
            return (row,alpha(col)+1)
        else:
            return (row,(alpha(col[0])+1)*26+alpha(col[1])+1)



def special_char_remove(string, cell=None):
    if string == None:
        string = ''
    string = str(string)
    # I think the following characters are produced from the text editor converted our UTF-8 
    #   characters into some other character set, like ISO-8859-1.
    for char in _global.special_char:
        string = string.replace(char, _global.special_char[char])
    for char in _global.extra_special_char:
        string = string.replace(char, _global.extra_special_char[char])

    string = ftfy.fix_text(string)

    # For some reaseon ftfy.fix_text(.) messes with some of the characters we replaced
    #  however, the first time (above) is needed for the ftfy to work.
    for char in _global.special_char:
        string = string.replace(char, _global.special_char[char])
    for char in _global.extra_special_char:
        string = string.replace(char, _global.extra_special_char[char])



    for i in ['Ã','¥','¶','â','Â','¼','½','¾']:
        if i in string:
            with open('zzz_REVIEW_STRING.txt','a+') as f:
                try:
                    f.write(f"REVIEW STRING in {_global.excelName} {cell}:\n {string}\n\n")
                except:
                    print(f"REVIEW STRING in {_global.excelName} {cell}:\n {string}\n\n")
            break
    return string


def ensure_encryption(wb):
    for ws_ in wb.sheetnames:
        ws = wb[ws_]
        for row in range(1,ws.max_row+1):
            for col in ['A','J','AL','IV','IW']:
                col = xl_to_num(col)
                string = ws.cell(row=row, column=col).value
                ws.cell(row=row, column=col).value = special_char_remove(string,(row,col))
    return wb



# Trys to find the workbook
def excel_open(excelName):
    try:
        return xl.load_workbook('{}.xlsx'.format(excelName))#,formatting_info=True,on_demand=True)
    except:
        raise Exception("Spreadsheet {} doesn't exists or is not of extension XLSX...".format(excelName))


def cell_read(loc, workbook):
    row,col = xl_to_num(loc)
    return workbook.cell(row=row,column=col).value

def xl_to_alpha(num):
    first = 0
    # Alphabet will be one indexed, that is, A = 1 = chr(1+64).
    while num-26>0:
        first += 1
        num -= 26
    if first == 0:
        return f'{chr(num+64)}'
    else:
        return f'{chr(first+64)}{chr(num+64)}'




def filterName(OrigName):
    dct = {'first':'','last':'','middle':'','suffix':'','email':'','institution':''}

    tmpName = OrigName.split(', ')
    dct['last'] = tmpName[0]
    first_middle = tmpName[1]

    fix,dot = False, False
    for i in first_middle:
        if i == '.':
            dot = True
        if i == '(' and dot:
            fix = True
            break
    if fix:
        nname = ''
        start = False
        for i in first_middle:
            if i == '(':
                start = True
                continue
            if i == ')':
                break
            if start:
                nname += i
                continue
        first_middle = nname
    tmp = first_middle.split(' ')
    dct['first'] = tmp[0]
    if len(tmp) > 1:
        dct['middle'] = ' '.join(tmp[1:])

    nums = [str(i) for i in range(0,10)]
    if len(tmpName) > 2:
        for i in range(2,len(tmpName)+1):
            for j in tmpName[i]:
                if j in nums:
                    break
            else:
                dct['suffix'] = tmpName[i]
                return dct
    return dct


def email_search(name):
    dct = filterName(name.lower())
    try:
        return _global.Authority_Lookup_Dict[dct['last']][dct['first']][dct['middle']]
    except KeyError:
        return ''



def recordAuthor(srow,scol,name):
    _global.rdsheet.cell(row = srow, column = scol).value = _global.authorDict[name]['first']
    _global.rdsheet.cell(row = srow, column = scol+1).value = _global.authorDict[name]['middle']
    _global.rdsheet.cell(row = srow, column = scol+2).value = _global.authorDict[name]['last']
    _global.rdsheet.cell(row = srow, column = scol+3).value = _global.authorDict[name]['suffix']
    _global.rdsheet.cell(row = srow, column = scol+4).value = _global.authorDict[name]['email']
    _global.rdsheet.cell(row = srow, column = scol+5).value = _global.authorDict[name]['institution']


def dictPrint(dct):
    print('First: {}\tMiddle: {}\tLast: {}\n\tSuffix: {}\tEmail: {}\tInstitution: {}\n'.format(
        dct['first'],
        dct['middle'],
        dct['last'],
        dct['suffix'],
        dct['email'],
        dct['institution']))



def main():    
    errors = []
    for i in range(2,_global.rdsheet.max_row+1):
        try:
            # Count MST authors
            authors = cell_read(f'H:{i}',_global.rdsheet)
            if authors == None:
                authors = ''
            _global.rdsheet.cell(row = i, column = xl_to_num('E')).value = len(authors.split('<br'))
            #Pulling names from authors
            authors = special_char_remove(cell_read(f'{_global.author_column}:{i}',_global.rdsheet)).split(' and ')
            # Record number of authors
            if _global.rdsheet.cell(row = i, column = xl_to_num('D')).value in [None,'',' ']:
                _global.rdsheet.cell(row = i, column = xl_to_num('D')).value = len(authors)
        except AttributeError:
            print(f'Check Excel format for {excelName}, make sure authors are in correct column.')
            return
        for name in sorted(authors):
            try:
                dctName = filterName(name)
            except:
                errors.append('ERROR IN ROW: {}\tAuthor: {}'.format(i,name))
                continue
            try:
                _global.authorDict[name]
            except KeyError:
                dctName['email'] = email_search(name)
                if dctName['email']:
                    if dctName['email'] == '...':
                        dctName['email'] = ''
                    dctName['institution'] = 'Missouri University of Science and Technology'
                _global.authorDict[name] = dctName
        if len(errors) == 0:
            srow,scol = xl_to_num(f'{xl_to_alpha(xl_to_num(author_column)+1)}:{i}') #Column of First Name should be one to the right of the Author column.
            for num,name in zip(range(0,len(authors)),authors):
                if num > 27:
                    break
                recordAuthor(srow,(scol)+num*7,name)
        _global.auth_counter = 3

    if len(errors) == 0:
        print('\n\n\n\n')
        for author in _global.authorDict:
            dictPrint(_global.authorDict[author])
        print('\n\n\n')
        _global.rb = ensure_encryption(_global.rb)
        _global.rb.save('{}_Complete.xlsx'.format(excelName))

    else:
        print('ERROR IN {}'.format(excelName))
        for error in errors:
            print('\t{}'.format(error))


if __name__ == "__main__":
    try:
        _global.authority_control_lookup = xl.load_workbook(_global.authority_control_lookup_path,read_only=True)
        _global.authority_control_lookup = _global.authority_control_lookup[_global.authority_control_lookup.sheetnames[0]]
    except:
        raise Exception("Spreadsheet '{}' doesn't exists or is not of extension XLSX...".format(authority_control_lookup_path))
    
    while True:
        print('\n\n\n\n')
        path = 'R:/storage/libarchive/a/Student Processing/0.5. Author split names'
        os.chdir(path)
        pot = []
        complete = ''
        try:
            os.remove('zzz_REVIEW_STRING.txt')
        except:
            pass
        for i in os.listdir():
            name, ext = os.path.splitext(i)
            if '_Complete' in i:
                complete += name

        for i in os.listdir():
            name, ext = os.path.splitext(i)
            if 'xl' in ext and '~' not in i and name not in complete:
                print(f'{len(pot)}. {name}')
                pot.append(name)
            elif 'null' in ext.lower():
                tmp = name.split('_')
                _global.author_column = tmp[-1]

        try:
            select_index = int(input('Selection Number: '))
        except:
            print('\n\nEnding Program...\n\n')
            quit()


        _global.excelName = pot[select_index]
        _global.rb = excel_open(_global.excelName)
        _global.rdsheet = _global.rb[rb.sheetnames[0]]

        main()