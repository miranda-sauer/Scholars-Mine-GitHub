import os
import ftfy
import xlrd as rd
import xlwt as wt
import openpyxl as xl
from xlutils.copy import copy as xlcopy




def special_char_remove(string):
    if string == None:
        string = ''
    string = str(string)
    #import ftfy
    # I think the following characters are produced from the text editor converted our UTF-8 
    #   characters into some other character set, like ISO-8859-1.
    dct = {'â€œ': '“',
            'â€': '”',
            'â€™': '’',
            'â€˜': '‘',
            'â€”': '–',
            'â€“': '—',
            'â€¢': '-',
            'â€¦': '…',
            'â€…':' ',
            '/&':'&',
            '/%':'%',
            'Â°':'°',
            'Ã—':'x',
            'ÅŸ':'ş'}
    extra_special = {'Ã‚':''}
    for char in dct:
        string = string.replace(char, dct[char])
    for char in extra_special:
        string = string.replace(char, extra_special[char])
    string = ftfy.fix_text(string)
    return string

def filter_special_char_from_workbook(wb):
    for ws_ in wb.sheetnames:
        ws = wb[ws_]
        for row in range(1,ws.max_row+1):
            for col in [1,10,38,257]:
                string = ws.cell(row=row, column=col).value
                ws.cell(row=row, column=col).value = special_char_remove(string)
    return wb





def filterName(OrigName):
    dct = {'first':'','last':'','middle':'','suffix':'','email':'','institution':''}

    tmpName = OrigName.split(', ')
    dct['last'] = tmpName[0]
    first_middle = tmpName[1]

    fix,dot = False, False
    for i in first_middle:
        if i == '.':
            dot = True
        if i == '(' and dot:
            fix = True
            break
    if fix:
        nname = ''
        start = False
        for i in first_middle:
            if i == '(':
                start = True
                continue
            if i == ')':
                break
            if start:
                nname += i
                continue
        first_middle = nname
    tmp = first_middle.split(' ')
    dct['first'] = tmp[0]
    if len(tmp) > 1:
        dct['middle'] = ' '.join(tmp[1:])

    nums = [str(i) for i in range(0,10)]
    if len(tmpName) > 2:
        for i in range(2,len(tmpName)+1):
            for j in tmpName[i]:
                if j in nums:
                    break
            else:
                dct['suffix'] = tmpName[i]
                return dct
    return dct



# What xlutils.copy is to .xls, this is the .xlsx
def copy_XLS_(tmpwb):
    finalwb = xl.Workbook()
    for x in finalwb.get_sheet_names(): finalwb.remove_sheet(finalwb.get_sheet_by_name(x))
    for sheet in tmpwb.get_sheet_names():
        tmpsheet = tmpwb[sheet]
        finalsheet = finalwb.create_sheet(title = sheet)
        for i in range(tmpsheet.max_row):
            for j in range(tmpsheet.max_column):
                finalsheet.cell(row = i+1, column = j+1, value = tmpsheet.cell(row=i+1,column=j+1).value) # row, column, value
    return finalwb

# Trys to find the workbook
def excel_open(excelName):
    try:
        return xl.load_workbook('{}.xlsm'.format(excelName))#,formatting_info=True,on_demand=True)
    except:
        try:
            return xl.load_workbook('{}.xlsx'.format(excelName))#,formatting_info=True,on_demand=True)
        except:
            raise Exception("Spreadsheet {} doesn't exists...".format(excelName))


# Converts a character to it's index in alphabet where A = 0 and Z = 25
def alpha(letter):
    return ord(letter)-65


# Give excel location such as 'A5' and returns a (row, column) tuple 'A5' -> (5,0)
def cell_call(loc):
    col,row = loc.split(':')
    row = int(row) - 1
    if len(col) == 1:
        return (row,alpha(col))
    else:
        return (row,(alpha(col[0])+1)*26+alpha(col[1]))




authority_control_lookup = rd.open_workbook('R:/storage/libarchive/a/Student Processing/0.5. Author split names/CODE/Authority Control_Lookups - Faculty.xls').sheet_by_index(0)




# Looks in Authority Control spreadsheet for our author, if they are not our author '' is returned and 
# they are marked as False in authorDict for future reading.
def email_search(name):
    dct = filterName(name)
    for i in dct:
        if dct[i] == False:
            dct[i] = ''

    # authority_control_lookup is a global workbook declared at the top!

    for i in range(1,authority_control_lookup.nrows+1):
        if cell_read_('P:{}'.format(i),authority_control_lookup) == dct['last']:
            if cell_read_('N:{}'.format(i),authority_control_lookup) == dct['first']:
                if cell_read_('O:{}'.format(i),authority_control_lookup) == dct['middle']:
                    if cell_read_('R:{}'.format(i),authority_control_lookup) == '':
                        return '...'
                    else:
                        return cell_read_('R:{}'.format(i),authority_control_lookup)
    return ''

authorDict = {}
def recordAuthor(srow,scol,name):
    dct = authorDict[name]
    w_sheet.cell(row = srow, column = scol, value = dct['first'])
    w_sheet.cell(row = srow, column = scol+1, value = dct['middle'])
    w_sheet.cell(row = srow, column = scol+2, value = dct['last'])
    w_sheet.cell(row = srow, column = scol+3, value = dct['suffix'])
    w_sheet.cell(row = srow, column = scol+4, value = dct['email'])
    w_sheet.cell(row = srow, column = scol+5, value = dct['institution'])


def dictPrint(dct):
    print('First: {}\tMiddle: {}\tLast: {}\n\tSuffix: {}\tEmail: {}\tInstitution: {}\n'.format(
        dct['first'],
        dct['middle'],
        dct['last'],
        dct['suffix'],
        dct['email'],
        dct['institution']))





def main():
    global wb
    errors = []
    for i in range(2,rdsheet.max_row+1):
        try:
            auth_col = author_column
            authors = cell_read(f'{auth_col}:{i}').split(' and ') #Pulling names from authors
        except AttributeError:
            print(f'Check Excel format for {excelName}, make sure authors are in correct column.')
            return
        for name in authors:
            try:
                dctName = filterName(name)
            except:
                errors.append('ERROR IN ROW: {}\tAuthor: {}'.format(i,name))
                continue
            try:
                authorDict[name]
            except KeyError:
                dctName['email'] = email_search(name)
                if dctName['email']:
                    if dctName['email'] == '...':
                        dctName['email'] = ''
                    dctName['institution'] = 'Missouri University of Science and Technology'
                authorDict[name] = dctName
        if len(errors) == 0:
            srow,scol = cell_call('BK:{}'.format(i)) #Column of First Name
            for num,name in zip(range(0,len(authors)),authors):
                if num > 27:
                    break
                recordAuthor(srow+1,(scol+1)+num*7,name)

    if len(errors) == 0:
        print('\n\n\n\n')
        for author in authorDict:
            dictPrint(authorDict[author])
        print('\n\n\n')
        wb = filter_special_char_from_workbook(wb)
        wb.save('{}_Complete.xlsx'.format(excelName))

    else:
        print('ERROR IN {}'.format(excelName))
        for error in errors:
            print('\t{}'.format(error))




if __name__ == '__main__':
    while True:

        print('\n\n\n\n')
        path = 'R:/storage/libarchive/a/Student Processing/0.5. Author split names'#str(input("Path to Spreadsheet (current directory = '.'): "))
        os.chdir(path)
        author_column = ''
        pot = []
        complete = ''
        for i in os.listdir():
            name, ext = os.path.splitext(i)
            if '_Complete' in i:
                complete += name

        for i in os.listdir():
            name, ext = os.path.splitext(i)
            if 'xl' in ext and '~' not in i and name not in complete:
                print(f'{len(pot)}. {name}')
                pot.append(name)
            elif 'txt' in ext:
                tmp = name.split('_')
                author_column = tmp[-1]

        try:
            select_index = int(input('Selection Number: '))
        except:
            input('Ending Program...')
            break


        excelName = pot[select_index]
        rb = excel_open(excelName)
        rdsheet = rb[rb.get_sheet_names()[0]]

        wb = copy_XLS_(rb)
        w_sheet = wb[wb.get_sheet_names()[0]]

        def cell_read(loc,worksheet=rdsheet):
            row,col = cell_call(loc)
            try:
                return worksheet.cell(rowx=row,colx=col)
            except TypeError:
                return worksheet[loc.replace(':','')].value

        def cell_read_(loc,workbook=rdsheet):
            row,col = cell_call(loc)
            return workbook.cell_value(rowx=row,colx=col)

        main()

