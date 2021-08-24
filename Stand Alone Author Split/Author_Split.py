# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           IMPORT                                                                #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
import os
from __init__ import fix_text
import sqlite3
import pickle
import openpyxl as xl

import sys
_global = sys.modules[__name__] # Allows access to 'global' variables defined below

from tkinter import *
from tkinter import filedialog
from tkinter import ttk
from tkinter import messagebox
import time
import datetime



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           GLOBAL VARIABLES                                                        #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

rdsheet = None
author_column = ''
excelName = ''

authority_database = sqlite3.connect('R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Stand Alone Author Split/faculty.db')
authority_cursor = authority_database.cursor()

authorDict = {}
rb = None

special_char = pickle.load(open('R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Stand Alone Author Split/special_char.pickle','rb'))
extra_special_char = pickle.load(open('R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Stand Alone Author Split/extra_special_char.pickle','rb'))



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           FUNCTIONS                                                               #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

# Converts a character to it's index in alphabet where A = 0 and Z = 25
def alpha(letter):
    return ord(letter)-65

# Give excel location such as 'A5' and returns a (row, column) tuple 'A5' -> (5,0)
def xl_to_num(loc):
    loc = loc.split(':')
    if len(loc) == 1:
        col = loc[0]
        if len(col) == 1:
            return alpha(col)+1
        else:
            return (alpha(col[0])+1)*26+alpha(col[1])+1
    else:
        col,row = loc
        row = int(row)
        if len(col) == 1:
            return (row,alpha(col)+1)
        else:
            return (row,(alpha(col[0])+1)*26+alpha(col[1])+1)



def special_char_remove(string, cell=None):
    if string == None:
        string = ''
    string = str(string)
    # I think the following characters are produced from the text editor converted our UTF-8 
    #   characters into some other character set, like ISO-8859-1.
    for char in _global.special_char:
        string = string.replace(char, _global.special_char[char])
    for char in _global.extra_special_char:
        string = string.replace(char, _global.extra_special_char[char])

    string = fix_text(string)

    # For some reaseon ftfy.fix_text(.) messes with some of the characters we replaced
    #  however, the first time (above) is needed for the ftfy to work.
    for char in _global.special_char:
        string = string.replace(char, _global.special_char[char])
    for char in _global.extra_special_char:
        string = string.replace(char, _global.extra_special_char[char])



    for i in ['Ã','¥','¶','â','Â','¼','½','¾']:
        if i in string:
            with open('zzz_REVIEW_STRING.txt','a+') as f:
                try:
                    f.write(f"REVIEW STRING in {_global.excelName} {cell}:\n {string}\n\n")
                except:
                    print(f"REVIEW STRING in {_global.excelName} {cell}:\n {string}\n\n")
            break
    return string


def ensure_encryption(wb):
    for ws_ in wb.sheetnames:
        ws = wb[ws_]
        for row in range(1,ws.max_row+1):
            for col in ['A','J','AL','IV','IW']:
                col = xl_to_num(col)
                string = ws.cell(row=row, column=col).value
                ws.cell(row=row, column=col).value = special_char_remove(string,(row,col))
    return wb



# Trys to find the workbook
def excel_open(excelName):
    try:
        return xl.load_workbook('{}'.format(excelName))#,formatting_info=True,on_demand=True)
    except:
        raise Exception("Spreadsheet {} doesn't exists or is not of extension XLSX...".format(excelName))

def cell_read(loc, workbook):
    row, col = xl_to_num(loc)
    return workbook.cell(row = row, column = col).value

def xl_to_alpha(num):
    first = 0
    # Alphabet will be one indexed, that is, A = 1 = chr(1+64).
    while num-26>0:
        first += 1
        num -= 26
    if first == 0:
        return f'{chr(num+64)}'
    else:
        return f'{chr(first+64)}{chr(num+64)}'

def filterName(OrigName):
    dct = {'first':'','last':'','middle':'','suffix':'','email':'','institution':''}

    tmpName = OrigName.split(', ')
    dct['last'] = tmpName[0]
    first_middle = tmpName[1]

    fix,dot = False, False
    for i in first_middle:
        if i == '.':
            dot = True
        if i == '(' and dot:
            fix = True
            break
    if fix:
        nname = ''
        start = False
        for i in first_middle:
            if i == '(':
                start = True
                continue
            if i == ')':
                break
            if start:
                nname += i
                continue
        first_middle = nname
    tmp = first_middle.split(' ')
    dct['first'] = tmp[0]
    if len(tmp) > 1:
        dct['middle'] = ' '.join(tmp[1:])

    nums = [str(i) for i in range(0,10)]
    if len(tmpName) > 2:
        for i in range(2,len(tmpName)+1):
            for j in tmpName[i]:
                if j in nums:
                    break
            else:
                dct['suffix'] = tmpName[i]
                return dct
    for key in dct:
        if dct[key] == '':
            dct[key] = None
    return dct

def email_search(name):
    dct = filterName(name.lower())
    tmp = authority_cursor.execute('SELECT email FROM faculty WHERE (last_name COLLATE NOCASE) = :last AND (first_name COLLATE NOCASE) = :first', dct).fetchall()
    if len(tmp) > 1:
        tmp = authority_cursor.execute('SELECT email FROM faculty WHERE (last_name COLLATE NOCASE) = :last AND (first_name COLLATE NOCASE) = :first AND (middle_name COLLATE NOCASE) =:middle', dct).fetchall()
    if len(tmp) == 1:
        return tmp[0][0]
    return ''

def recordAuthor(srow,scol,name):
    _global.rdsheet.cell(row = srow, column = scol).value = _global.authorDict[name]['first']
    _global.rdsheet.cell(row = srow, column = scol+1).value = _global.authorDict[name]['middle']
    _global.rdsheet.cell(row = srow, column = scol+2).value = _global.authorDict[name]['last']
    _global.rdsheet.cell(row = srow, column = scol+3).value = _global.authorDict[name]['suffix']
    _global.rdsheet.cell(row = srow, column = scol+4).value = _global.authorDict[name]['email']
    _global.rdsheet.cell(row = srow, column = scol+5).value = _global.authorDict[name]['institution']

def dictPrint(dct):
    print('First: {}\tMiddle: {}\tLast: {}\n\tSuffix: {}\tEmail: {}\tInstitution: {}\n'.format(
        dct['first'],
        dct['middle'],
        dct['last'],
        dct['suffix'],
        dct['email'],
        dct['institution']))



# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #
#           DISPLAY WINDOW                                                          #
# * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * * #

def open_author_split():
    # Create window
    author_split = Tk()
    author_split.title("Author Split Program")
    author_split.configure(bg = '#003B49')
    author_split.iconbitmap('R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Stand Alone Author Split/S&T_Logo.ico')

    # Center the window on the screen
    window_w = 400 # window width
    window_h = 290 # window height

    screen_w = author_split.winfo_screenwidth() # screen width
    screen_h = author_split.winfo_screenheight() # screen height

    x_coor = (screen_w/2) - (window_w/2) #middle of screen x coordinate
    y_coor = (screen_h/2) #lower half of screen y coordinate

    author_split.geometry("%dx%d+%d+%d" % (window_w, window_h, x_coor, y_coor)) # place in middle

    #Create Progress Bar
    progress = ttk.Progressbar(author_split, orient = HORIZONTAL, length = 370, mode = 'determinate')

    # Create a frame
    frame = LabelFrame(author_split, text = "Select an Input File", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')

    # Create a label
    task = Label(author_split, text = "Waiting for a file", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')
    file_label = Label(author_split, text = "File: N/A", bg = '#003B49', fg = 'white', font = 'tungsten 12 bold')

    # Open Help Function
    def open_help():
        # Open word document
        try:
            os.startfile("R:/storage/libarchive/b/1. Processing/8. Other Projects/Scholars-Mine-GitHub/Control Panel/Documentation/Author Split Help.docx")
        except:
            error_popup("Could not open help file.")

    # Create a button
    help_button = Button(author_split, text = "Help", command = lambda : open_help(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    exit_button = Button(author_split, text = "Exit Author Split", command = author_split.destroy, bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")

    # Place in window
    frame.pack(padx = 20, pady = 10)
    file_label.pack(padx = 20, pady = (0, 20))
    progress.pack(padx = 5, pady = 5)
    task.pack(padx = 20)
    help_button.pack(padx = (331, 10), pady = 0)
    exit_button.pack(padx = (242, 10), pady = (5, 0))

    # Error Message Popup
    def error_popup(error_message):
        messagebox.showerror("Error", error_message)

    # Warning Message Popup
    def warning_popup(warning_message):
        messagebox.showwarning("Warning", warning_message)

    # Update Progress Bar
    def update_progress(p, t):  
        # Update bar value
        progress['value'] = (p/2)*100

        # Update bar label
        task.config(text = t)
        task.pack() 

        # Refresh window (very important line!)
        author_split.update_idletasks()
        time.sleep(0.2)

    # Function Click
    def browse():
        # Reset progress bar
        update_progress(0, "Waiting for a file")

        # Open file
        author_split.filename = filedialog.askopenfilename(initialdir = "R:/storage/libarchive/b/1. Processing/8. Other Projects/Excel Files", title = "Select Input", filetypes = (("Excel Workbook", "*.xlsx"),))
        if author_split.filename == "":
            warning_popup("No folder selected.")
            file_label.config(text = "Folder: N/A")
            del author_split.filename
        else:
            #Get file name
            name = author_split.filename
            for i in range(len(name)):
                if "/" in name[-(i)]:
                    name = "File: " + name[-(i-1):]
                    break

            #Update Folder Name Label
            file_label.config(text = name)

    # Main Function
    def main():
        update_progress(1, "Splitting Authors...")

#       max_col = len([c for c in _global.rdsheet.iter_cols(min_row = 1, max_row = 1, values_only = True) if c[0] is not None])
        max_col = 100 # make bigger if critical columns have an index higher than 100

        # Gets column index for the column titled "author"
        def author_index():
            # search all column headers
            for col in range(1, max_col + 1):               
                col_head = _global.rdsheet.cell(1, col).value
                # looks for column header that is "author"
                if 'author' in col_head and '_' not in col_head:
                    # returns index
                    return col

        # Gets column index for the column titled "total_author_count"
        def total_author_count_index():
            # search all column headers
            for col in range(1, max_col + 1):               
                col_head = _global.rdsheet.cell(1, col).value
                # looks for column header that is "total_author_count"
                if 'total_author_count' in col_head:
                    # returns index
                    return col

        # Gets column index for the column titled "faculty_author_count"
        def faculty_author_count_index():
            # search all column headers
            for col in range(1, max_col + 1):               
                col_head = _global.rdsheet.cell(1, col).value
                # looks for column header that is "faculty_author_count"
                if 'faculty_author_count' in col_head:
                    # returns index
                    return col

        # Gets column index for the column titled "authorized_name"
        def authorized_name_index():
            # search all column headers
            for col in range(1, max_col + 1):               
                col_head = _global.rdsheet.cell(1, col).value
                # looks for column header that is "authorized_name"
                if 'authorized_name' in col_head:
                    # returns index
                    return col

        _global.excelName = author_split.filename       # Assign excel file name
        _global.rb = excel_open(_global.excelName)      # Open excel file
        _global.rdsheet = _global.rb[rb.sheetnames[0]]  # Access excel file sheet
        
        errors = []
        for i in range(2,_global.rdsheet.max_row+1):
            try:
                # Count MST authors
                authors = _global.rdsheet.cell(i, authorized_name_index()).value

                if authors == None:
                    authors = ''

                _global.rdsheet.cell(i, faculty_author_count_index()).value = len(authors.split('<br'))
                    
                #Pulling names from authors
                authors = special_char_remove(_global.rdsheet.cell(i, author_index()).value).split(' and ')
                    
                # Record number of authors
                if _global.rdsheet.cell(i, total_author_count_index()).value in [None,'',' ']:
                    _global.rdsheet.cell(i, total_author_count_index()).value = len(authors)
                    
            except AttributeError:
                print(f'Check Excel format for {excelName}, make sure authors are in correct column.')
                return
                
            for name in sorted(authors):
                try:
                    dctName = filterName(name)
                except:
                    errors.append('ERROR IN ROW: {}\tAuthor: {}'.format(i,name))
                    continue
                try:
                    _global.authorDict[name]
                except KeyError:
                    dctName['email'] = email_search(name)
                    if dctName['email']:
                        if dctName['email'] == '...':
                            dctName['email'] = ''
                        dctName['institution'] = 'Missouri University of Science and Technology'
                    _global.authorDict[name] = dctName
                
            if len(errors) == 0:
                srow,scol = xl_to_num(f'{xl_to_alpha(author_index()+1)}:{i}') #Column of First Name should be one to the right of the Author column.
                for num,name in zip(range(0,len(authors)),authors):
                    if num > 27:
                        break
                    recordAuthor(srow,(scol)+num*7,name)
            _global.auth_counter = 3
        if len(errors) == 0:
            print('\n\n\n\n')
            for author in _global.authorDict:
                dictPrint(_global.authorDict[author])
            print('\n\n\n')
            _global.rb = ensure_encryption(_global.rb)
            _global.rb.save('{}_Complete.xlsx'.format((excelName)[:len(excelName) - 25]))
        else:
            print('ERROR IN {}'.format(excelName))
            for error in errors:
                print('\t{}'.format(error))

        update_progress(2, "Excel Created")

    # Start Button Function
    def start():
        # Run program and update progress bar
        try:
            main()
        except AttributeError:
            error_popup("No folder selected. Browse to select a folder.")
        except:
            error_popup("There was an unknown error, the file could not be processed.")
    
    # Create a button
    browse_button = Button(frame, text = "Browse", command = lambda : browse(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    start_button = Button(frame, text = "Start", command = lambda : start(), bg = '#78BE20', fg = '#003B49', font = 'tungsten 12 bold', borderwidth = 1, relief = "ridge")
    
    # Place in frame
    browse_button.grid(row = 0, column = 0, padx = (15, 0), pady = 15)
    start_button.grid(row = 0, column = 1, padx = 15, pady = 15)

    # Keeps window open until closed
    author_split.mainloop()
