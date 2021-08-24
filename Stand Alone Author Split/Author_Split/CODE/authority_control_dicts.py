import pickle
import openpyxl as xl


# Converts a character to it's index in alphabet where A = 1 and Z = 26
def alpha(letter):
    return ord(letter)-64

# Give excel location such as 'A5' and returns a (row, column) tuple 'A5' -> (5,0)
def to_digit(loc):
    col,row = loc.split(':')
    row = int(row)
    if len(col) == 1:
        return (row,alpha(col)+1)
    else:
        return (row,(alpha(col[0])+1)*26+alpha(col[1])+1)

def scramble(string):
    for i in [' ',',','_', '-', '.pdf','.','0','1','2','3','4','5','6','7','8','9']:
        string = string.replace(i,'')
    new_string = ''.join(sorted(list(string)))
    return new_string

sensitive_dct = {}
lower_dct = {}
wb = xl.load_workbook('R:/storage/libarchive/a/Student Processing/zzz_Program_Dependencies/Authority_Control_Lookups.xlsx',read_only=True)
ws = wb[wb.sheetnames[0]]
row = 0
max_row = ws.max_row+1
try:
    for row in range(2,max_row):
        print(f"{row} of {max_row}")
        key = []
        key += [ws.cell(row = row, column = alpha('P')).value]
        key += [ws.cell(row = row, column = alpha('N')).value]
        key += [ws.cell(row = row, column = alpha('O')).value]
        key += [ws.cell(row = row, column = alpha('R')).value]
        for num, el in enumerate(key):
            if el == None and num != 3:
                key[num] = ''
            elif el==None and num==3:
                key[num] = '...'
        key = [x.lower() for x in key]
        try:
            sensitive_dct[key[0]]
            lower_dct[key[0].lower()]
        except:
            sensitive_dct[key[0]] = {}
            lower_dct[key[0].lower()] = {}
        try:
            sensitive_dct[key[0]][key[1]]
            lower_dct[key[0].lower()][key[1].lower()]
        except:
            sensitive_dct[key[0]][key[1]] = {}
            lower_dct[key[0].lower()][key[1].lower()] = {}
        sensitive_dct[key[0]][key[1]][key[2]] = key[3]
        lower_dct[key[0].lower()][key[1].lower()][key[2].lower()] = key[3]
    
    with open('Authority_Control_Dict_Case_Sensitive.pickle','wb') as f:
        pickle.dump(sensitive_dct, f, pickle.HIGHEST_PROTOCOL)
    with open('Authority_Control_Dict_lowercase.pickle','wb') as f:
        pickle.dump(lower_dct, f, pickle.HIGHEST_PROTOCOL)

except:
    print(f'FAIL at row {row}')
    input("Exiting...")